"""
Daily update: 2026-05-07
Creates v1.24 from v1.23.

TBU retries (May 6 games):
- Match 56: BUF/MTL G1 → Buffalo 4, Montreal 2 (CONFIRMED)
- Match 57: VGK/ANA G2 → Vegas 3, Anaheim 1 (CONFIRMED)

New TBU (tonight, May 7):
- Match 58: CAR/PHI G3 → TBU (8:00 PM ET tonight, no recap yet)

G7 empty slots for R1 series that ended in 6 games or fewer (Matches 45-48, 50) remain
empty — those games were never played.

Bracket: D27 VGK/ANA badge 1→2, N11 BUF/MTL badge 0→1,
         N7 BUF GREEN→YELLOW, N15 MTL GREEN→WHITE.
"""
import shutil
from copy import copy
import openpyxl
from openpyxl.styles import PatternFill, Font

REPO = '/home/user/nhl-2026-playoffs'
src = f'{REPO}/2026 NHL Playoffs_v1.23.xlsx'
dst = f'{REPO}/2026 NHL Playoffs_v1.24.xlsx'
shutil.copyfile(src, dst)

wb = openpyxl.load_workbook(dst)
ws = wb['2026 NHL Playoffs_By Dates']

# ── Style palette ──────────────────────────────────────────────────────────
WHITE      = PatternFill('solid', fgColor='FFFFFF')
GREEN      = PatternFill('solid', fgColor='92D050')
YELLOW     = PatternFill('solid', fgColor='FFFF00')
GREY       = PatternFill('solid', fgColor='D9D9D9')
TAN        = PatternFill('solid', fgColor='FFF2CC')
BLACK_BOLD = Font(name='Calibri', size=11, bold=True,  color='000000')

# ── By Dates updates ───────────────────────────────────────────────────────
# row: (result, scorers)
updates = {
    # Match 56 (row 57): BUF/MTL G1, May 6 — previously TBU
    # Goals in order:
    #   P1 4:31  J. Doan (EV)         → 1-0 BUF
    #   P1 13:26 R. McLeod (PP)       → 2-0 BUF
    #   P1 19:16 N. Suzuki (PP) MTL   → 2-1
    #   P2  3:32 J. Greenway (EV, GW) → 3-1 BUF  [GW: 3rd BUF goal > MTL final 2]
    #   P2  9:01 B. Byram (PP)        → 4-1 BUF
    #   P2 16:31 K. Dach (EV) MTL     → 4-2  [confirmed via TSN: "3:29 remaining P2"]
    #   P3: scoreless (TSN confirmed BUF led 4-2 after P2; no P3 goals)
    57: (
        'Buffalo 4, Montreal 2',
        'BUF: J. Doan, R. McLeod (PP), J. Greenway (GW), B. Byram (PP) | MTL: N. Suzuki (PP), K. Dach'
    ),

    # Match 57 (row 58): VGK/ANA G2, May 6 — previously TBU
    # Goals in order:
    #   P2 ~3:14 B. Howden (EV)             → 1-0 VGK
    #   P3  6:03 M. Granlund (EV) ANA        → 1-1
    #   P3 15:02 I. Barbashev (EV, GW)       → 2-1 VGK  [non-icing controversy; GW]
    #   P3 19:54 M. Marner (EN)              → 3-1 VGK  [190-foot EN, 6 sec left]
    58: (
        'Vegas 3, Anaheim 1',
        'VGK: B. Howden, I. Barbashev (GW), M. Marner (EN) | ANA: M. Granlund'
    ),

    # Match 58 (row 59): CAR/PHI G3, May 7 (tonight, 8:00 PM ET) — mark TBU
    59: ('TBU', 'TBU'),
}

for row, (result, scorers) in updates.items():
    ws.cell(row=row, column=9).value  = result
    ws.cell(row=row, column=10).value = scorers

# ── Bracket updates ────────────────────────────────────────────────────────
b = wb['Bracket']

# Update R2 series-score badges
# VGK/ANA: was VGK 1 - 0 ANA → now VGK 2 - 0 ANA (after G2 win)
b['D27'].value = 'VGK 2 - 0 ANA'
b['D27'].fill  = TAN
b['D27'].font  = BLACK_BOLD

# BUF/MTL: was BUF 0 - 0 MTL → now BUF 1 - 0 MTL (after G1 BUF win)
b['N11'].value = 'BUF 1 - 0 MTL'
b['N11'].fill  = TAN
b['N11'].font  = BLACK_BOLD

# Update R2 team cell fills to reflect actual series status
# BUF leads 1-0: BUF → YELLOW (leading), MTL → WHITE (trailing)
b['N7'].fill  = YELLOW
b['N7'].font  = BLACK_BOLD
b['N15'].fill = WHITE
b['N15'].font = BLACK_BOLD

# VGK leads 2-0: VGK already YELLOW, ANA already WHITE — no change needed
# CAR leads 2-0 (G3 tonight, TBU): CAR already YELLOW, PHI already WHITE — no change
# COL leads 2-0 (G3 Sat May 9, future): COL already YELLOW, MIN already WHITE — no change

# ── Save ───────────────────────────────────────────────────────────────────
wb.save(dst)
print(f'Saved: {dst}')
