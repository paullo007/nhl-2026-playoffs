import openpyxl
import shutil
from openpyxl.styles import PatternFill, Font

src = '/home/user/nhl-2026-playoffs/2026 NHL Playoffs_v1.49.xlsx'
dst = '/home/user/nhl-2026-playoffs/2026 NHL Playoffs_v1.50.xlsx'
shutil.copyfile(src, dst)

wb = openpyxl.load_workbook(dst)
ws = wb['2026 NHL Playoffs_By Dates']

# === Row 96: Match 95, SCF Game 3, CAR @ VGK, Jun 6, 2026 ===
# Result: Vegas 5, Carolina 4 (2OT) — VGK leads series 2-1
# VGK goals: Hertl (PP, 2nd period), Marner x3 (hat trick, 2nd period), Theodore (2OT winner)
# CAR goals: Martinook (7:02 3rd), Hall (7:29 3rd), Staal (7:42 3rd, 39-sec burst), Svechnikov PP (1:42 left)
ws.cell(row=96, column=9).value = 'Vegas 5, Carolina 4 (2OT)'
ws.cell(row=96, column=10).value = (
    'VGK: T. Hertl (PP), M. Marner (3), S. Theodore (2OT winner) | '
    'CAR: J. Martinook, T. Hall, J. Staal, A. Svechnikov (PP)'
)

# === Bracket updates ===
b = wb['Bracket']

WHITE  = PatternFill('solid', fgColor='FFFFFF')
YELLOW = PatternFill('solid', fgColor='FFFF00')
BLACK_BOLD = Font(name='Calibri', size=11, bold=True, color='000000')

# SCF series badge: VGK now leads 2-1
b['I18'].value = 'STANLEY\nCUP\nFINAL\nVGK 2 - 1 CAR'

# VGK team cell (H19): leading — apply YELLOW
b['H19'].value = 'Vegas Golden Knights\n(leads SCF 2-1)'
b['H19'].fill = YELLOW
b['H19'].font = BLACK_BOLD

# CAR team cell (J19): trailing — WHITE (default)
b['J19'].value = 'Carolina Hurricanes\n(trails SCF 2-1)'
b['J19'].fill = WHITE
b['J19'].font = BLACK_BOLD

wb.save(dst)
print(f'Saved: {dst}')
print('\nUpdated cells:')
print(f'  Row 96 col I: {ws.cell(row=96, column=9).value}')
print(f'  Row 96 col J: {ws.cell(row=96, column=10).value}')
print(f'  Bracket I18:  {b["I18"].value!r}')
print(f'  Bracket H19:  {b["H19"].value!r}')
print(f'  Bracket J19:  {b["J19"].value!r}')
