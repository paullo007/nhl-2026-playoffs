import openpyxl
import shutil
from openpyxl.styles import PatternFill, Font

REPO = '/home/user/nhl-2026-playoffs'
src = f'{REPO}/2026 NHL Playoffs_v1.29.xlsx'
dst = f'{REPO}/2026 NHL Playoffs_v1.30.xlsx'
shutil.copyfile(src, dst)

wb = openpyxl.load_workbook(dst)
ws = wb['2026 NHL Playoffs_By Dates']
b  = wb['Bracket']

# === BY DATES UPDATE — May 16 2026 run ===
#
# Row 73 (M#72): VGK @ ANA, Game 6 of West R2, actually played May 14
# VGK wins series 4-2 with a 5-1 win
# Goal log (chronological):
#   P1  1:02 — M. Marner (VGK) EV -> 1-0 VGK
#   P1 11:30 — B. Howden (VGK) SH -> 2-0 VGK (3rd SH goal of playoffs; GW = puts VGK at loser+1)
#   P1 17:19 — S. Theodore (VGK) PP -> 3-0 VGK
#   P2 12:46 — M. Granlund (ANA) PP -> 3-1 (snap shot from left face-off circle; T. Terry assist)
#   P3  2:52 — P. Dorofeyev (VGK) EV -> 4-1 VGK (feed from Barbashev after stealing Carlson's clear)
#   P3 13:32 — P. Dorofeyev (VGK) EV -> 5-1 VGK (short side from above goal line; 6:28 remaining)
# GW: Howden — VGK's 2nd goal (ANA final = 1; VGK's 2nd = ANA+1) ✓
# Count: VGK 5 (Marner + Howden + Theodore + Dorofeyev×2); ANA 1 (Granlund) ✓
# Series: VGK wins 4-2
ws.cell(row=73, column=9).value  = 'Vegas 5, Anaheim 1 (VGK wins series 4-2)'
ws.cell(row=73, column=10).value = 'VGK: M. Marner, B. Howden (SH, GW), S. Theodore (PP), P. Dorofeyev (2) | ANA: M. Granlund (PP)'

# Row 75 (M#74): BUF @ MTL, Game 6 of East R2, May 16 (8:00 PM ET Bell Centre)
# Recap not yet available — game tonight, set TBU for retry tomorrow
ws.cell(row=75, column=9).value  = 'TBU'
ws.cell(row=75, column=10).value = 'TBU'

# === BRACKET UPDATES ===
# VGK wins R2 4-2: VGK advances to Western Conference Final vs COL

WHITE      = PatternFill('solid', fgColor='FFFFFF')
GREEN      = PatternFill('solid', fgColor='92D050')
YELLOW     = PatternFill('solid', fgColor='FFFF00')
TAN        = PatternFill('solid', fgColor='FFF2CC')
GREY       = PatternFill('solid', fgColor='D9D9D9')

BLACK_BOLD = Font(name='Calibri', size=11, bold=True, color='000000')
GREY_TEXT  = Font(name='Calibri', size=11, bold=True, color='808080', strike=True)

# D27 badge: VGK 3 - 2 ANA -> VGK 4 - 2 ANA
b['D27'].value = 'VGK 4 - 2 ANA'
b['D27'].fill  = TAN
b['D27'].font  = BLACK_BOLD

# D23 (VGK R2 cell): YELLOW -> GREEN (clinched)
b['D23'].fill = GREEN
b['D23'].font = BLACK_BOLD

# D31 (ANA R2 cell): WHITE -> GREY with strikethrough (eliminated)
b['D31'].fill = GREY
b['D31'].font = GREY_TEXT

# F27: Pacific Final placeholder -> VGK advances, with opponent (COL) now known
b['F27'].value = 'Vegas Golden Knights\n(advanced — vs. Colorado Avalanche)'
b['F27'].fill  = GREEN
b['F27'].font  = BLACK_BOLD

# F11: COL was "opponent TBD" -> update opponent now that VGK has advanced
b['F11'].value = 'Colorado Avalanche\n(advanced — vs. Vegas Golden Knights)'
b['F11'].fill  = GREEN
b['F11'].font  = BLACK_BOLD

wb.save(dst)
print(f'Saved: {dst}')
