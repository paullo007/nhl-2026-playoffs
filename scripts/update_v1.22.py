import openpyxl
import shutil
from openpyxl.styles import PatternFill, Font

REPO = '/home/user/nhl-2026-playoffs'
src = f'{REPO}/2026 NHL Playoffs_v1.21.xlsx'
dst = f'{REPO}/2026 NHL Playoffs_v1.22.xlsx'
shutil.copyfile(src, dst)

wb = openpyxl.load_workbook(dst)
ws = wb['2026 NHL Playoffs_By Dates']

# Row 50: Match 49 · TBL/MTL Game 7 · May 3 · previously TBU → now confirmed
# Final: Montreal 2, Tampa Bay 1 (MTL wins series 4-3)
# Scoring:
#   P1 18:39 - N. Suzuki (MTL) — EV, off Guhle shot deflection
#   P2 13:27 - D. James (TBL) — PP goal (TBL 1-for-2 on PP)
#   P3 11:07 - A. Newhook (MTL) — GW (Newhook batted in rebound off Vasilevskiy's back)
ws.cell(row=50, column=9).value = 'Montreal 2, Tampa Bay 1 (MTL wins series 4-3)'
ws.cell(row=50, column=10).value = 'MTL: N. Suzuki, A. Newhook (GW) | TBL: D. James (PP)'

# --- Bracket updates ---
b = wb['Bracket']

GREEN      = PatternFill('solid', fgColor='92D050')
GREY       = PatternFill('solid', fgColor='D9D9D9')
WHITE      = PatternFill('solid', fgColor='FFFFFF')
BLACK_BOLD = Font(name='Calibri', size=11, bold=True, color='000000')
GREY_TEXT  = Font(name='Calibri', size=11, bold=True, color='808080', strike=True)

# Series badge P15: TBL 3 - 3 MTL → TBL 3 - 4 MTL (MTL wins)
b['P15'].value = 'TBL 3 - 4 MTL'

# P13 (TBL): WHITE → GREY + strikethrough (eliminated)
b['P13'].fill = GREY
b['P13'].font = GREY_TEXT

# P17 (MTL): WHITE → GREEN (advanced / clinched series)
b['P17'].fill = GREEN
b['P17'].font = BLACK_BOLD

# N15: MTL advances; opponent is Buffalo (BUF already in N7)
n15 = b['N15']
n15.value = 'Montreal Canadiens\n(advanced — vs. Buffalo Sabres)'
n15.fill = GREEN
n15.font = BLACK_BOLD

# N7: BUF opponent now confirmed as MTL
n7 = b['N7']
n7.value = 'Buffalo Sabres\n(advanced — vs. Montreal Canadiens)'
n7.fill = GREEN
n7.font = BLACK_BOLD

wb.save(dst)
print('Saved:', dst)
