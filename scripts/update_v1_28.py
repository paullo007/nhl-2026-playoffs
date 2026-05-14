import openpyxl
import shutil
from openpyxl.styles import PatternFill, Font

REPO = '/home/user/nhl-2026-playoffs'
src = f'{REPO}/2026 NHL Playoffs_v1.27.xlsx'
dst = f'{REPO}/2026 NHL Playoffs_v1.28.xlsx'
shutil.copyfile(src, dst)

wb = openpyxl.load_workbook(dst)
ws = wb['2026 NHL Playoffs_By Dates']
b  = wb['Bracket']

# === BY DATES UPDATES — May 14 2026 run ===
#
# Row 68 (M#67): CAR/PHI G5 — series over (CAR won 4-0), game never played. Leave None.
#
# Row 69 (M#68): VGK/ANA G5, May 12 (workbook had TBD Thu May14, game was actually Tue May12)
# VGK 3, ANA 2 (OT)
# Goals (chronological):
#   P1 12:36 — B. Sennecke (ANA) PP -> 1-0 ANA
#   P1 16:13 — P. Dorofeyev (VGK) PP -> 1-1 (stole puck in offensive zone, slot shot)
#   P3  4:38 — T. Hertl (VGK) EV -> 2-1 VGK (recovered loose puck, five-hole)
#   P3 ~16:55 — O. Zellweger (ANA) EV -> 2-2 (bar down from left dot, 3:05 remaining)
#   OT  4:10 — P. Dorofeyev (VGK) OT winner -> 3-2 VGK (rebound near left post off Eichel shot)
# GW: Dorofeyev OT goal (VGK's 3rd goal = 1 more than ANA's final total of 2)
# Count: VGK 3 (Dorofeyev 2 + Hertl 1 = 3); ANA 2 (Sennecke + Zellweger = 2)
# VGK leads series 3-2
ws.cell(row=69, column=9).value  = 'Vegas 3, Anaheim 2 (OT)'
ws.cell(row=69, column=10).value = 'VGK: P. Dorofeyev (2, 1 PP, OT winner), T. Hertl | ANA: B. Sennecke (PP), O. Zellweger'

# Row 70 (M#69): COL/MIN G5, May 13 (workbook had TBD Thu May14, game was actually Wed May13)
# COL 4, MIN 3 (OT) — COL wins series 4-1, advances to Western Conference Final
# Goals (chronological):
#   P1  0:34 — M. Johansson (MIN) EV -> 1-0 MIN (one-timer from left circle)
#   P1 11:03 — N. Foligno (MIN) EV -> 2-0 MIN (snuck around Kulak, redirected Sturm pass five-hole)
#   P1 15:56 — N. Foligno (MIN) EV -> 3-0 MIN (tap-in off Sturm rebound across crease)
#   P2 11:00 — P. Kelly (COL) EV -> 3-1 MIN (deflected Burns' point shot short side; confirmed EV)
#   P3 16:27 — J. Drury (COL) EV -> 3-2 MIN (deflected Toews' point shot in high slot)
#   P3 18:37 — N. MacKinnon (COL) EV -> 3-3 (wrist shot top shelf, COL goalie pulled 6-on-5)
#   OT  3:52 — B. Kulak (COL) OT winner -> 4-3 COL (one-timer off Necas crossing pass, right dot)
# GW: Kulak (OT winner)
# Count: COL 4 (Kelly+Drury+MacKinnon+Kulak=4); MIN 3 (Johansson+Foligno*2=3)
# COL wins series 4-1
ws.cell(row=70, column=9).value  = 'Colorado 4, Minnesota 3 (OT) (COL wins series 4-1)'
ws.cell(row=70, column=10).value = 'COL: P. Kelly, J. Drury, N. MacKinnon, B. Kulak (OT winner) | MIN: M. Johansson, N. Foligno (2)'

# Row 71 (M#70): BUF/MTL G5, May 14 — 7 PM ET tonight, not yet played
ws.cell(row=71, column=9).value  = 'TBU'
ws.cell(row=71, column=10).value = 'TBU'

# === BRACKET UPDATES ===
# Series state after this run (non-TBU only):
#   COL/MIN: COL wins 4-1 (G5 confirmed)
#   VGK/ANA: VGK leads 3-2 (G5 confirmed)
#   BUF/MTL: Tied 2-2 (G5 TBU — do not update badge)
#   CAR/PHI: CAR wins 4-0 (already done in v1.26/v1.27)

WHITE      = PatternFill('solid', fgColor='FFFFFF')
YELLOW     = PatternFill('solid', fgColor='FFFF00')
GREEN      = PatternFill('solid', fgColor='92D050')
GREY       = PatternFill('solid', fgColor='D9D9D9')

BLACK_BOLD = Font(name='Calibri', size=11, bold=True, color='000000')
GREY_TEXT  = Font(name='Calibri', size=11, bold=True, color='808080', strike=True)

# --- COL/MIN: COL wins series 4-1 ---
b['D11'].value = 'COL 4 - 1 MIN'
b['D7'].fill   = GREEN      # COL: clinched (was YELLOW leading)
b['D7'].font   = BLACK_BOLD
b['D15'].fill  = GREY       # MIN: eliminated (was WHITE trailing)
b['D15'].font  = GREY_TEXT  # strikethrough for eliminated team

# COL advances to Western Conference Central Final (cell F11)
b['F11'].value = 'Colorado Avalanche\n(advanced — opponent TBD)'
b['F11'].fill  = GREEN
b['F11'].font  = BLACK_BOLD

# --- VGK/ANA: VGK leads 3-2 ---
b['D27'].value = 'VGK 3 - 2 ANA'
b['D23'].fill  = YELLOW     # VGK: leading (was WHITE tied)
b['D23'].font  = BLACK_BOLD
b['D31'].fill  = WHITE      # ANA: trailing (was WHITE tied, stays WHITE)
b['D31'].font  = BLACK_BOLD

# --- BUF/MTL: Tied 2-2 (G5 TBU — no badge change) ---
# N11 stays 'BUF 2 - 2 MTL'; N7 and N15 fills unchanged

wb.save(dst)
print(f'Saved: {dst}')
