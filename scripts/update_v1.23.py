"""
Daily update: 2026-05-06
Creates v1.23 from v1.22.
- Adds Round 2 rows to By Dates sheet (Match 51–78, rows 52–79)
- Fills confirmed R2 results (Match 51–55)
- Marks TBU for tonight's games (Match 56–57: BUF/MTL G1, VGK/ANA G2)
- Leaves future games empty
- Adds R2 series-score badges at D11, D27, N11, N27
- Updates R2 team cell fill colors to reflect series status
"""
import shutil
from copy import copy
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

REPO = '/home/user/nhl-2026-playoffs'
src = f'{REPO}/2026 NHL Playoffs_v1.22.xlsx'
dst = f'{REPO}/2026 NHL Playoffs_v1.23.xlsx'
shutil.copyfile(src, dst)

wb = openpyxl.load_workbook(dst)
ws = wb['2026 NHL Playoffs_By Dates']

# ── Style palette ──────────────────────────────────────────────────────────
WHITE      = PatternFill('solid', fgColor='FFFFFF')
GREEN      = PatternFill('solid', fgColor='92D050')
YELLOW     = PatternFill('solid', fgColor='FFFF00')
GREY       = PatternFill('solid', fgColor='D9D9D9')
TAN        = PatternFill('solid', fgColor='FFF2CC')
BLACK_BOLD = Font(name='Calibri', size=11, bold=True,  color='000000')
BLACK_NRM  = Font(name='Calibri', size=11, bold=False, color='000000')
GREY_TEXT  = Font(name='Calibri', size=11, bold=True,  color='808080', strike=True)

# ── Copy row styles from reference row 51 ─────────────────────────────────
ref_row = 51
ref_styles = {}
for col in range(1, 11):
    cell = ws.cell(row=ref_row, column=col)
    ref_styles[col] = {
        'fill':          copy(cell.fill),
        'font':          copy(cell.font),
        'alignment':     copy(cell.alignment),
        'border':        copy(cell.border),
        'number_format': cell.number_format,
    }
ref_height = ws.row_dimensions[ref_row].height or 36

# ── Round 2 row data ───────────────────────────────────────────────────────
# Columns: match#, round, series, game, matchup, date_local, date_sgt, arena,
#          result (None=empty/future, 'TBU'=tonight), scorers
R2 = [
    # ─── Match 51: CAR/PHI G1 · Sat May 2 · at CAR ── CONFIRMED ─────────
    (51, 'Round 2', 'Carolina Hurricanes vs Philadelphia Flyers', 'Game 1',
     'Philadelphia Flyers vs Carolina Hurricanes (at Carolina Hurricanes)',
     '8:00 PM ET, Sat, May2', '8:00 AM SGT, Sun, May3',
     'Lenovo Center, Raleigh, NC',
     'Carolina 3, Philadelphia 0',
     'CAR: L. Stankoven (2, incl. GW), J. Blake | PHI: (shutout)'),

    # ─── Match 52: COL/MIN G1 · Sun May 3 · at COL ── CONFIRMED ─────────
    (52, 'Round 2', 'Colorado Avalanche vs Minnesota Wild', 'Game 1',
     'Minnesota Wild vs Colorado Avalanche (at Colorado Avalanche)',
     '9:00 PM ET, Sun, May3', '9:00 AM SGT, Mon, May4',
     'Ball Arena, Denver, CO',
     'Colorado 9, Minnesota 6',
     'COL: S. Malinski, J. Drury, A. Lehkonen, N. Blankenburg, D. Toews, '
     'C. Makar (2), N. Kadri (GW), N. MacKinnon (EN) | '
     'MIN: R. Hartman, M. Johansson, V. Tarasenko, M. Foligno (SH), '
     'M. Zuccarello, +1 (unconfirmed)'),

    # ─── Match 53: CAR/PHI G2 · Mon May 4 · at CAR ── CONFIRMED ─────────
    (53, 'Round 2', 'Carolina Hurricanes vs Philadelphia Flyers', 'Game 2',
     'Philadelphia Flyers vs Carolina Hurricanes (at Carolina Hurricanes)',
     '7:00 PM ET, Mon, May4', '7:00 AM SGT, Tue, May5',
     'Lenovo Center, Raleigh, NC',
     'Carolina 3, Philadelphia 2 (OT)',
     'CAR: N. Ehlers (PP), S. Jarvis, T. Hall (OT winner) | '
     'PHI: J. Drysdale (PP), S. Couturier'),

    # ─── Match 54: VGK/ANA G1 · Mon May 4 · at VGK ── CONFIRMED ─────────
    (54, 'Round 2', 'Vegas Golden Knights vs Anaheim Ducks', 'Game 1',
     'Anaheim Ducks vs Vegas Golden Knights (at Vegas Golden Knights)',
     '6:30 PM PT, Mon, May4', '9:30 AM SGT, Tue, May5',
     'T-Mobile Arena, Paradise, NV',
     'Vegas 3, Anaheim 1',
     'VGK: B. Howden, I. Barbashev (GW), M. Marner | ANA: M. Granlund'),

    # ─── Match 55: COL/MIN G2 · Tue May 5 · at COL ── CONFIRMED ─────────
    (55, 'Round 2', 'Colorado Avalanche vs Minnesota Wild', 'Game 2',
     'Minnesota Wild vs Colorado Avalanche (at Colorado Avalanche)',
     '8:00 PM ET, Tue, May5', '8:00 AM SGT, Wed, May6',
     'Ball Arena, Denver, CO',
     'Colorado 5, Minnesota 2',
     'COL: M. Necas, G. Landeskog (PP), N. Roy (GW), N. MacKinnon (PP), '
     'V. Nichushkin (EN) | MIN: K. Kaprizov, M. Johansson'),

    # ─── Match 56: BUF/MTL G1 · Wed May 6 · at BUF ── TBU (tonight) ─────
    (56, 'Round 2', 'Buffalo Sabres vs Montreal Canadiens', 'Game 1',
     'Montreal Canadiens vs Buffalo Sabres (at Buffalo Sabres)',
     '7:00 PM ET, Wed, May6', '7:00 AM SGT, Thu, May7',
     'KeyBank Center, Buffalo, NY',
     'TBU', 'TBU'),

    # ─── Match 57: VGK/ANA G2 · Wed May 6 · at VGK ── TBU (tonight) ─────
    (57, 'Round 2', 'Vegas Golden Knights vs Anaheim Ducks', 'Game 2',
     'Anaheim Ducks vs Vegas Golden Knights (at Vegas Golden Knights)',
     '9:30 PM ET, Wed, May6', '9:30 AM SGT, Thu, May7',
     'T-Mobile Arena, Paradise, NV',
     'TBU', 'TBU'),

    # ─── Match 58: CAR/PHI G3 · Thu May 7 · at PHI ── FUTURE ────────────
    (58, 'Round 2', 'Carolina Hurricanes vs Philadelphia Flyers', 'Game 3',
     'Carolina Hurricanes vs Philadelphia Flyers (at Philadelphia Flyers)',
     '8:00 PM ET, Thu, May7', '8:00 AM SGT, Fri, May8',
     'Wells Fargo Center, Philadelphia, PA',
     None, None),

    # ─── Match 59: BUF/MTL G2 · Fri May 8 · at BUF ── FUTURE ────────────
    (59, 'Round 2', 'Buffalo Sabres vs Montreal Canadiens', 'Game 2',
     'Montreal Canadiens vs Buffalo Sabres (at Buffalo Sabres)',
     'TBD, Fri, May8', 'TBD, Sat, May9',
     'KeyBank Center, Buffalo, NY',
     None, None),

    # ─── Match 60: VGK/ANA G3 · TBD · at ANA ── FUTURE ──────────────────
    (60, 'Round 2', 'Vegas Golden Knights vs Anaheim Ducks', 'Game 3',
     'Vegas Golden Knights vs Anaheim Ducks (at Anaheim Ducks)',
     'TBD, Fri, May8', 'TBD, Sat, May9',
     'Honda Center, Anaheim, CA',
     None, None),

    # ─── Match 61: COL/MIN G3 · Sat May 9 · at MIN ── FUTURE ────────────
    (61, 'Round 2', 'Colorado Avalanche vs Minnesota Wild', 'Game 3',
     'Colorado Avalanche vs Minnesota Wild (at Minnesota Wild)',
     '9:00 PM ET, Sat, May9', '9:00 AM SGT, Sun, May10',
     'Xcel Energy Center, St. Paul, MN',
     None, None),

    # ─── Match 62: BUF/MTL G3 · Sun May 10 · at MTL ── FUTURE ──────────
    (62, 'Round 2', 'Buffalo Sabres vs Montreal Canadiens', 'Game 3',
     'Buffalo Sabres vs Montreal Canadiens (at Montreal Canadiens)',
     '7:00 PM ET, Sun, May10', '7:00 AM SGT, Mon, May11',
     'Bell Centre, Montreal, QC',
     None, None),

    # ─── Match 63: CAR/PHI G4 · TBD · at PHI ── FUTURE ──────────────────
    (63, 'Round 2', 'Carolina Hurricanes vs Philadelphia Flyers', 'Game 4',
     'Carolina Hurricanes vs Philadelphia Flyers (at Philadelphia Flyers)',
     'TBD, Mon, May11', 'TBD, Tue, May12',
     'Wells Fargo Center, Philadelphia, PA',
     None, None),

    # ─── Match 64: VGK/ANA G4 · TBD · at ANA ── FUTURE ──────────────────
    (64, 'Round 2', 'Vegas Golden Knights vs Anaheim Ducks', 'Game 4',
     'Vegas Golden Knights vs Anaheim Ducks (at Anaheim Ducks)',
     'TBD, Mon, May11', 'TBD, Tue, May12',
     'Honda Center, Anaheim, CA',
     None, None),

    # ─── Match 65: COL/MIN G4 · Mon May 11 · at MIN ── FUTURE ───────────
    (65, 'Round 2', 'Colorado Avalanche vs Minnesota Wild', 'Game 4',
     'Colorado Avalanche vs Minnesota Wild (at Minnesota Wild)',
     'TBD, Mon, May11', 'TBD, Tue, May12',
     'Xcel Energy Center, St. Paul, MN',
     None, None),

    # ─── Match 66: BUF/MTL G4 · Tue May 12 · at MTL ── FUTURE ──────────
    (66, 'Round 2', 'Buffalo Sabres vs Montreal Canadiens', 'Game 4',
     'Buffalo Sabres vs Montreal Canadiens (at Montreal Canadiens)',
     'TBD, Tue, May12', 'TBD, Wed, May13',
     'Bell Centre, Montreal, QC',
     None, None),

    # ─── Match 67: CAR/PHI G5 · TBD · at CAR ── FUTURE ──────────────────
    (67, 'Round 2', 'Carolina Hurricanes vs Philadelphia Flyers', 'Game 5',
     'Philadelphia Flyers vs Carolina Hurricanes (at Carolina Hurricanes)',
     'TBD, Thu, May14', 'TBD, Fri, May15',
     'Lenovo Center, Raleigh, NC',
     None, None),

    # ─── Match 68: VGK/ANA G5 · TBD · at VGK ── FUTURE ──────────────────
    (68, 'Round 2', 'Vegas Golden Knights vs Anaheim Ducks', 'Game 5',
     'Anaheim Ducks vs Vegas Golden Knights (at Vegas Golden Knights)',
     'TBD, Thu, May14', 'TBD, Fri, May15',
     'T-Mobile Arena, Paradise, NV',
     None, None),

    # ─── Match 69: COL/MIN G5 · TBD · at COL ── FUTURE ──────────────────
    (69, 'Round 2', 'Colorado Avalanche vs Minnesota Wild', 'Game 5',
     'Minnesota Wild vs Colorado Avalanche (at Colorado Avalanche)',
     'TBD, Thu, May14', 'TBD, Fri, May15',
     'Ball Arena, Denver, CO',
     None, None),

    # ─── Match 70: BUF/MTL G5 · Thu May 14 · at BUF ── FUTURE ──────────
    (70, 'Round 2', 'Buffalo Sabres vs Montreal Canadiens', 'Game 5',
     'Montreal Canadiens vs Buffalo Sabres (at Buffalo Sabres)',
     'TBD, Thu, May14', 'TBD, Fri, May15',
     'KeyBank Center, Buffalo, NY',
     None, None),

    # ─── Match 71: CAR/PHI G6 · TBD · at PHI ── FUTURE ──────────────────
    (71, 'Round 2', 'Carolina Hurricanes vs Philadelphia Flyers', 'Game 6',
     'Carolina Hurricanes vs Philadelphia Flyers (at Philadelphia Flyers)',
     'TBD, Sat, May16', 'TBD, Sun, May17',
     'Wells Fargo Center, Philadelphia, PA',
     None, None),

    # ─── Match 72: VGK/ANA G6 · TBD · at ANA ── FUTURE ──────────────────
    (72, 'Round 2', 'Vegas Golden Knights vs Anaheim Ducks', 'Game 6',
     'Vegas Golden Knights vs Anaheim Ducks (at Anaheim Ducks)',
     'TBD, Sat, May16', 'TBD, Sun, May17',
     'Honda Center, Anaheim, CA',
     None, None),

    # ─── Match 73: COL/MIN G6 · TBD · at MIN ── FUTURE ──────────────────
    (73, 'Round 2', 'Colorado Avalanche vs Minnesota Wild', 'Game 6',
     'Colorado Avalanche vs Minnesota Wild (at Minnesota Wild)',
     'TBD, Sat, May16', 'TBD, Sun, May17',
     'Xcel Energy Center, St. Paul, MN',
     None, None),

    # ─── Match 74: BUF/MTL G6 · Sat May 16 · at MTL ── FUTURE ──────────
    (74, 'Round 2', 'Buffalo Sabres vs Montreal Canadiens', 'Game 6',
     'Buffalo Sabres vs Montreal Canadiens (at Montreal Canadiens)',
     'TBD, Sat, May16', 'TBD, Sun, May17',
     'Bell Centre, Montreal, QC',
     None, None),

    # ─── Match 75: CAR/PHI G7 · TBD · at CAR ── FUTURE ──────────────────
    (75, 'Round 2', 'Carolina Hurricanes vs Philadelphia Flyers', 'Game 7',
     'Philadelphia Flyers vs Carolina Hurricanes (at Carolina Hurricanes)',
     'TBD, Mon, May18', 'TBD, Tue, May19',
     'Lenovo Center, Raleigh, NC',
     None, None),

    # ─── Match 76: VGK/ANA G7 · TBD · at VGK ── FUTURE ──────────────────
    (76, 'Round 2', 'Vegas Golden Knights vs Anaheim Ducks', 'Game 7',
     'Anaheim Ducks vs Vegas Golden Knights (at Vegas Golden Knights)',
     'TBD, Mon, May18', 'TBD, Tue, May19',
     'T-Mobile Arena, Paradise, NV',
     None, None),

    # ─── Match 77: COL/MIN G7 · TBD · at COL ── FUTURE ──────────────────
    (77, 'Round 2', 'Colorado Avalanche vs Minnesota Wild', 'Game 7',
     'Minnesota Wild vs Colorado Avalanche (at Colorado Avalanche)',
     'TBD, Mon, May18', 'TBD, Tue, May19',
     'Ball Arena, Denver, CO',
     None, None),

    # ─── Match 78: BUF/MTL G7 · Mon May 18 · at BUF ── FUTURE ──────────
    (78, 'Round 2', 'Buffalo Sabres vs Montreal Canadiens', 'Game 7',
     'Montreal Canadiens vs Buffalo Sabres (at Buffalo Sabres)',
     'TBD, Mon, May18', 'TBD, Tue, May19',
     'KeyBank Center, Buffalo, NY',
     None, None),
]

start_row = 52
for idx, row_data in enumerate(R2):
    r = start_row + idx
    match_num, rnd, series, game, matchup, dl, ds, arena, result, scorers = row_data

    ws.cell(row=r, column=1).value = match_num
    ws.cell(row=r, column=2).value = rnd
    ws.cell(row=r, column=3).value = series
    ws.cell(row=r, column=4).value = game
    ws.cell(row=r, column=5).value = matchup
    ws.cell(row=r, column=6).value = dl
    ws.cell(row=r, column=7).value = ds
    ws.cell(row=r, column=8).value = arena
    ws.cell(row=r, column=9).value  = result  if result  is not None else None
    ws.cell(row=r, column=10).value = scorers if scorers is not None else None

    # Apply reference styles
    for col in range(1, 11):
        cell = ws.cell(row=r, column=col)
        s = ref_styles[col]
        cell.fill          = copy(s['fill'])
        cell.font          = copy(s['font'])
        cell.alignment     = copy(s['alignment'])
        cell.border        = copy(s['border'])
        cell.number_format = s['number_format']

    ws.row_dimensions[r].height = ref_height

print(f'Added {len(R2)} Round 2 rows (rows {start_row}–{start_row+len(R2)-1})')

# ── BRACKET UPDATES ────────────────────────────────────────────────────────
b = wb['Bracket']

# R2 series-score badges (new cells)
# Position: D11 (COL/MIN), D27 (VGK/ANA), N11 (BUF/MTL), N27 (CAR/PHI)
r2_badges = {
    'D11': 'COL 2 - 0 MIN',   # COL leads 2-0
    'D27': 'VGK 1 - 0 ANA',   # VGK leads 1-0
    'N11': 'BUF 0 - 0 MTL',   # Series not yet started (tonight)
    'N27': 'CAR 2 - 0 PHI',   # CAR leads 2-0
}
for coord, val in r2_badges.items():
    c = b[coord]
    c.value = val
    c.fill  = TAN
    c.font  = BLACK_BOLD

# R2 team cell fills — reflect series status
# COL leads 2-0: COL → YELLOW, MIN → WHITE
b['D7'].fill  = YELLOW   # COL leading
b['D7'].font  = BLACK_BOLD
b['D15'].fill = WHITE    # MIN trailing
b['D15'].font = BLACK_BOLD

# VGK leads 1-0: VGK → YELLOW, ANA → WHITE
b['D23'].fill = YELLOW   # VGK leading
b['D23'].font = BLACK_BOLD
b['D31'].fill = WHITE    # ANA trailing
b['D31'].font = BLACK_BOLD

# BUF/MTL G1 is tonight — series 0-0; keep both GREEN (series starts tonight)
# N7 and N15 already GREEN — no change needed

# CAR leads 2-0: CAR → YELLOW, PHI → WHITE
b['N23'].fill = YELLOW   # CAR leading
b['N23'].font = BLACK_BOLD
b['N31'].fill = WHITE    # PHI trailing
b['N31'].font = BLACK_BOLD

print('Bracket R2 badges added at D11, D27, N11, N27')
print('Bracket R2 team cell fills updated: COL/VGK/CAR → YELLOW, MIN/ANA/PHI → WHITE')

# ── SAVE ───────────────────────────────────────────────────────────────────
wb.save(dst)
print(f'Saved: {dst}')
