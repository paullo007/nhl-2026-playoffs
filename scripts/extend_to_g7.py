"""Extend the By Dates sheet with Game 5/6/7 rows for all in-progress series.
Skips swept series (CAR/OTT, COL/LAK).
Saves as v1.16, preserving v1.15 format conventions.
"""
import openpyxl
import shutil
from copy import copy

REPO = '/Users/paullo/01_PLO/02_CLAUDE CODE/02_NHL 2026 Playoffs'
src = f'{REPO}/2026 NHL Playoffs_v1.15.xlsx'
dst = f'{REPO}/2026 NHL Playoffs_v1.16.xlsx'
shutil.copyfile(src, dst)

wb = openpyxl.load_workbook(dst)
ws = wb['2026 NHL Playoffs_By Dates']

# New rows (chronological order). Format:
# (match_num, series_label, game, matchup, date_local, date_sgt, arena, result, scorers)
new_rows = [
    # Match 33 — PIT-PHI G5 (Mon Apr 27, played)
    (33, 'Round 1', 'Pittsburgh Penguins vs Philadelphia Flyers', 'Game 5',
     'Philadelphia Flyers vs Pittsburgh Penguins (at Pittsburgh Penguins)',
     '7:00 PM ET, Mon, Apr27', '7:00 AM SGT, Tue, Apr28',
     'PPG Paints Arena, Pittsburgh, PA',
     'Pittsburgh 3, Philadelphia 2',
     'PIT: E. Soderblom, C. Dewar, K. Letang (GW) | PHI: A. Bump, T. Sanheim'),

    # Match 34 — BUF-BOS G5 (Tue Apr 28)
    (34, 'Round 1', 'Buffalo Sabres vs Boston Bruins', 'Game 5',
     'Boston Bruins vs Buffalo Sabres (at Buffalo Sabres)',
     '7:30 PM ET, Tue, Apr28', '7:30 AM SGT, Wed, Apr29',
     'KeyBank Center, Buffalo, NY', None, None),

    # Match 35 — DAL-MIN G5 (Tue Apr 28)
    (35, 'Round 1', 'Dallas Stars vs Minnesota Wild', 'Game 5',
     'Minnesota Wild vs Dallas Stars (at Dallas Stars)',
     '7:00 PM CT, Tue, Apr28', '8:00 AM SGT, Wed, Apr29',
     'American Airlines Center, Dallas, TX', None, None),

    # Match 36 — EDM-ANA G5 (Tue Apr 28)
    (36, 'Round 1', 'Edmonton Oilers vs Anaheim Ducks', 'Game 5',
     'Anaheim Ducks vs Edmonton Oilers (at Edmonton Oilers)',
     '8:00 PM MT, Tue, Apr28', '10:00 AM SGT, Wed, Apr29',
     'Rogers Place, Edmonton, AB', None, None),

    # Match 37 — PIT-PHI G6 (Wed Apr 29)
    (37, 'Round 1', 'Pittsburgh Penguins vs Philadelphia Flyers', 'Game 6',
     'Pittsburgh Penguins vs Philadelphia Flyers (at Philadelphia Flyers)',
     '7:30 PM ET, Wed, Apr29', '7:30 AM SGT, Thu, Apr30',
     'Wells Fargo Center, Philadelphia, PA', None, None),

    # Match 38 — TBL-MTL G5 (Wed Apr 29)
    (38, 'Round 1', 'Tampa Bay Lightning vs Montreal Canadiens', 'Game 5',
     'Montreal Canadiens vs Tampa Bay Lightning (at Tampa Bay Lightning)',
     '7:00 PM ET, Wed, Apr29', '7:00 AM SGT, Thu, Apr30',
     'Amalie Arena, Tampa, FL', None, None),

    # Match 39 — VGK-UTA G5 (Wed Apr 29)
    (39, 'Round 1', 'Vegas Golden Knights vs Utah Mammoth', 'Game 5',
     'Utah Mammoth vs Vegas Golden Knights (at Vegas Golden Knights)',
     '7:00 PM PT, Wed, Apr29', '10:00 AM SGT, Thu, Apr30',
     'T-Mobile Arena, Paradise, NV', None, None),

    # Match 40 — EDM-ANA G6 (Thu Apr 30)
    (40, 'Round 1', 'Edmonton Oilers vs Anaheim Ducks', 'Game 6',
     'Edmonton Oilers vs Anaheim Ducks (at Anaheim Ducks)',
     'TBD, Thu, Apr30', 'TBD, Fri, May1',
     'Honda Center, Anaheim, CA', None, None),

    # Match 41 — DAL-MIN G6 (Thu Apr 30)
    (41, 'Round 1', 'Dallas Stars vs Minnesota Wild', 'Game 6',
     'Dallas Stars vs Minnesota Wild (at Minnesota Wild)',
     'TBD, Thu, Apr30', 'TBD, Fri, May1',
     'Xcel Energy Center, St. Paul, MN', None, None),

    # Match 42 — BUF-BOS G6 (Fri May 1)
    (42, 'Round 1', 'Buffalo Sabres vs Boston Bruins', 'Game 6',
     'Buffalo Sabres vs Boston Bruins (at Boston Bruins)',
     'TBD, Fri, May1', 'TBD, Sat, May2',
     'TD Garden, Boston, MA', None, None),

    # Match 43 — TBL-MTL G6 (Fri May 1)
    (43, 'Round 1', 'Tampa Bay Lightning vs Montreal Canadiens', 'Game 6',
     'Tampa Bay Lightning vs Montreal Canadiens (at Montreal Canadiens)',
     'TBD, Fri, May1', 'TBD, Sat, May2',
     'Bell Centre, Montreal, QC', None, None),

    # Match 44 — VGK-UTA G6 (Fri May 1)
    (44, 'Round 1', 'Vegas Golden Knights vs Utah Mammoth', 'Game 6',
     'Vegas Golden Knights vs Utah Mammoth (at Utah Mammoth)',
     'TBD, Fri, May1', 'TBD, Sat, May2',
     'Delta Center, Salt Lake City, UT', None, None),

    # Match 45 — PIT-PHI G7 (Sat May 2)
    (45, 'Round 1', 'Pittsburgh Penguins vs Philadelphia Flyers', 'Game 7',
     'Philadelphia Flyers vs Pittsburgh Penguins (at Pittsburgh Penguins)',
     'TBD, Sat, May2', 'TBD, Sun, May3',
     'PPG Paints Arena, Pittsburgh, PA', None, None),

    # Match 46 — EDM-ANA G7 (Sat May 2)
    (46, 'Round 1', 'Edmonton Oilers vs Anaheim Ducks', 'Game 7',
     'Anaheim Ducks vs Edmonton Oilers (at Edmonton Oilers)',
     'TBD, Sat, May2', 'TBD, Sun, May3',
     'Rogers Place, Edmonton, AB', None, None),

    # Match 47 — DAL-MIN G7 (Sat May 2)
    (47, 'Round 1', 'Dallas Stars vs Minnesota Wild', 'Game 7',
     'Minnesota Wild vs Dallas Stars (at Dallas Stars)',
     'TBD, Sat, May2', 'TBD, Sun, May3',
     'American Airlines Center, Dallas, TX', None, None),

    # Match 48 — BUF-BOS G7 (Sun May 3)
    (48, 'Round 1', 'Buffalo Sabres vs Boston Bruins', 'Game 7',
     'Boston Bruins vs Buffalo Sabres (at Buffalo Sabres)',
     'TBD, Sun, May3', 'TBD, Mon, May4',
     'KeyBank Center, Buffalo, NY', None, None),

    # Match 49 — TBL-MTL G7 (Sun May 3)
    (49, 'Round 1', 'Tampa Bay Lightning vs Montreal Canadiens', 'Game 7',
     'Montreal Canadiens vs Tampa Bay Lightning (at Tampa Bay Lightning)',
     'TBD, Sun, May3', 'TBD, Mon, May4',
     'Amalie Arena, Tampa, FL', None, None),

    # Match 50 — VGK-UTA G7 (Sun May 3)
    (50, 'Round 1', 'Vegas Golden Knights vs Utah Mammoth', 'Game 7',
     'Utah Mammoth vs Vegas Golden Knights (at Vegas Golden Knights)',
     'TBD, Sun, May3', 'TBD, Mon, May4',
     'T-Mobile Arena, Paradise, NV', None, None),
]

# Reference row to copy styling from (last existing row = 33)
ref_row = 33
start_row = 34

# Get reference styles per column
ref_styles = {}
for col in range(1, 11):
    cell = ws.cell(row=ref_row, column=col)
    ref_styles[col] = {
        'fill': copy(cell.fill),
        'font': copy(cell.font),
        'alignment': copy(cell.alignment),
        'border': copy(cell.border),
        'number_format': cell.number_format,
    }

# Set row height to match existing rows (typical 22)
existing_height = ws.row_dimensions[ref_row].height or 22

for idx, row_data in enumerate(new_rows):
    r = start_row + idx
    match_num, rd, series, game, matchup, dl, ds, arena, result, scorers = row_data

    ws.cell(row=r, column=1).value = match_num
    ws.cell(row=r, column=2).value = rd
    ws.cell(row=r, column=3).value = series
    ws.cell(row=r, column=4).value = game
    ws.cell(row=r, column=5).value = matchup
    ws.cell(row=r, column=6).value = dl
    ws.cell(row=r, column=7).value = ds
    ws.cell(row=r, column=8).value = arena
    if result is not None:
        ws.cell(row=r, column=9).value = result
    if scorers is not None:
        ws.cell(row=r, column=10).value = scorers

    # Apply reference styles to each cell
    for col in range(1, 11):
        cell = ws.cell(row=r, column=col)
        s = ref_styles[col]
        cell.fill = copy(s['fill'])
        cell.font = copy(s['font'])
        cell.alignment = copy(s['alignment'])
        cell.border = copy(s['border'])
        cell.number_format = s['number_format']

    ws.row_dimensions[r].height = existing_height

wb.save(dst)
print(f'Saved: {dst}')
print(f'Added {len(new_rows)} rows (rows {start_row}–{start_row + len(new_rows) - 1}, matches 33–50)')
print(f'Max row now: {ws.max_row}')
