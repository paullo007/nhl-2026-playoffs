import openpyxl
import shutil
from openpyxl.styles import PatternFill, Font

REPO = '/home/user/nhl-2026-playoffs'
src = f'{REPO}/2026 NHL Playoffs_v1.19.xlsx'
dst = f'{REPO}/2026 NHL Playoffs_v1.20.xlsx'
shutil.copyfile(src, dst)

wb = openpyxl.load_workbook(dst)
ws = wb['2026 NHL Playoffs_By Dates']

# May 1 confirmed games (previously TBU)
updates = {
    # Row 43: Match 42 · BUF @ BOS Game 6 · Buffalo 4, Boston 1 · BUF wins series 4-2
    # No PP goals (BUF 0-for-4 PP, BOS never had PP); Norris EN confirmed; GW = Samuelsson (2nd BUF goal, BOS ended with 1)
    43: ('Buffalo 4, Boston 1 (BUF wins series 4-2)',
         'BUF: A. Tuch, M. Samuelsson (GW), Z. Benson, J. Norris (EN) | BOS: D. Pastrnak'),

    # Row 44: Match 43 · TBL @ MTL Game 6 · Tampa Bay 1, Montreal 0 (OT) · Series tied 3-3 → Game 7
    # Goncalves scored off own rebound at 9:03 OT (EV); Dobes shutout (32 saves)
    44: ('Tampa Bay 1, Montreal 0 (OT)',
         'TBL: G. Goncalves (OT winner) | MTL: (shutout)'),

    # Row 45: Match 44 · VGK @ UTA Game 6 · Vegas 5, Utah 1 · VGK wins series 4-2
    # Howden EV (1st, 15:02); Marner EV GW (2nd, 19:15); Yamamoto EV (3rd, UTA cuts to 2-1);
    # Sissons EV (3rd, restores 3-1); Marner PP (3rd, 7:51 rem); Craig Smith EN (3rd, 3:36 rem)
    45: ('Vegas 5, Utah 1 (VGK wins series 4-2)',
         'VGK: B. Howden, M. Marner (GW), C. Sissons, M. Marner (PP), C. Smith (EN) | UTA: K. Yamamoto'),
}

for row, (result, scorers) in updates.items():
    ws.cell(row=row, column=9).value = result
    ws.cell(row=row, column=10).value = scorers

# --- Bracket updates ---
b = wb['Bracket']

GREEN      = PatternFill('solid', fgColor='92D050')
WHITE      = PatternFill('solid', fgColor='FFFFFF')
GREY       = PatternFill('solid', fgColor='D9D9D9')
BLACK_BOLD = Font(name='Calibri', size=11, bold=True, color='000000')
GREY_TEXT  = Font(name='Calibri', size=11, bold=True, color='808080', strike=True)

# Series score badges (format: TEAM_LEFT SCORE_LEFT - SCORE_RIGHT TEAM_RIGHT)
b['B23'].value = 'VGK 4 - 2 UTA'   # VGK clinched
b['P7'].value  = 'BUF 4 - 2 BOS'   # BUF clinched
b['P15'].value = 'TBL 3 - 3 MTL'   # forced to Game 7

# B21 (VGK): YELLOW → GREEN (series clinched, advanced)
b['B21'].fill = GREEN
b['B21'].font = BLACK_BOLD

# B25 (UTA): WHITE → GREY + strikethrough (eliminated)
b['B25'].fill = GREY
b['B25'].font = GREY_TEXT

# P5 (BUF): YELLOW → GREEN (series clinched, advanced)
b['P5'].fill = GREEN
b['P5'].font = BLACK_BOLD

# P9 (BOS): WHITE → GREY + strikethrough (eliminated)
b['P9'].fill = GREY
b['P9'].font = GREY_TEXT

# P17 (MTL): YELLOW → WHITE (series now tied 3-3, no longer leading)
b['P17'].fill = WHITE
b['P17'].font = BLACK_BOLD

# R2: D23 → VGK advances, faces ANA (confirmed opponent)
d23 = b['D23']
d23.value = 'Vegas Golden Knights\n(advanced — vs. Anaheim Ducks)'
d23.fill = GREEN
d23.font = BLACK_BOLD

# R2: D31 → ANA opponent now confirmed as VGK
d31 = b['D31']
d31.value = 'Anaheim Ducks\n(advanced — vs. Vegas Golden Knights)'
d31.fill = GREEN
d31.font = BLACK_BOLD

# R2: N7 → BUF advances; opponent TBD (TBL/MTL goes to Game 7 on May 3)
n7 = b['N7']
n7.value = 'Buffalo Sabres\n(advanced — opponent TBD)'
n7.fill = GREEN
n7.font = BLACK_BOLD

wb.save(dst)
print('Saved:', dst)
