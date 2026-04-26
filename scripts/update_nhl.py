import openpyxl
import shutil

src = '/Users/paullo/01_PLO/02_CLAUDE CODE/02_NHL 2026 Playoffs/2026 NHL Playoffs_v1.8.xlsx'
dst = '/Users/paullo/01_PLO/02_CLAUDE CODE/02_NHL 2026 Playoffs/2026 NHL Playoffs_v1.9.xlsx'
shutil.copyfile(src, dst)

wb = openpyxl.load_workbook(dst)

# --- Update "By Dates" sheet: Result (col I) and Goal Scorers (col J) ---
ws = wb['2026 NHL Playoffs_By Dates']

updates = {
    # Match 16 (row 17): Pittsburgh @ Philadelphia, Game 3, Apr 22
    17: ('Philadelphia 5, Pittsburgh 2',
         'PHI: T. Zegras, R. Ristolainen, N. Seeler, N. Cates, O. Tippett | PIT: E. Malkin, E. Karlsson'),
    # Match 17 (row 18): Dallas @ Minnesota, Game 3, Apr 22
    18: ('Dallas 4, Minnesota 3 (2OT)',
         'DAL: M. Rantanen (PP), J. Robertson, M. Duchene (PP), W. Johnston (2OT winner, PP) | MIN: M. Johansson (PP), J. Eriksson Ek, M. McCarron'),
    # Match 18 (row 19): Anaheim @ Edmonton, Game 2, Apr 22
    19: ('Anaheim 6, Edmonton 4',
         'ANA: C. Gauthier (2), R. Poehling (2, incl. EN), J. Trouba, A. Killorn (PP) | EDM: L. Draisaitl, C. Murphy, Z. Hyman, J. Samanski'),
    # Match 19 (row 20): Buffalo @ Boston, Game 3, Apr 23
    20: ('Buffalo 3, Boston 1',
         'BUF: B. Byram, A. Tuch (GW), N. Ostlund (EN) | BOS: T. Jeannot'),
    # Match 20 (row 21): Carolina @ Ottawa, Game 3, Apr 23
    21: ('Carolina 2, Ottawa 1',
         'CAR: L. Stankoven, J. Blake (GW) | OTT: D. Batherson'),
    # Match 21 (row 22): Colorado @ Los Angeles, Game 3, Apr 23
    22: ('Colorado 4, Los Angeles 2',
         'COL: G. Landeskog, C. Makar (GW), A. Lehkonen, B. Nelson | LAK: 2 goals (scorers unconfirmed)'),
    # Match 22 (row 23): Tampa Bay @ Montreal, Game 3, Apr 24
    23: ('Montreal 3, Tampa Bay 2 (OT)',
         'MTL: A. Texier, K. Dach, L. Hutson (OT winner) | TBL: B. Hagel, B. Point'),
    # Match 23 (row 24): Vegas @ Utah, Game 3, Apr 24
    24: ('Utah 4, Vegas 2',
         'UTA: L. Crouse (2), +2 (unconfirmed) | VGK: J. Eichel, N. Dowd'),
    # Match 24 (row 25): Edmonton @ Anaheim, Game 3, Apr 24
    25: ('Anaheim 7, Edmonton 4',
         'ANA: B. Sennecke, L. Carlsson, M. Granlund, J. LaCombe, +3 (unconfirmed) | EDM: V. Podkolzin, K. Kapanen, R. Nugent-Hopkins, +1 (unconfirmed)'),
    # Match 25 (row 26): Carolina @ Ottawa, Game 4, Apr 25 (3pm ET)
    26: ('Carolina 4, Ottawa 2 (CAR wins series 4-0)',
         'CAR: T. Hall, L. Stankoven (GW), S. Aho (2, both EN) | OTT: D. Batherson (PP), D. Cozens'),
    # Match 26 (row 27): Dallas @ Minnesota, Game 4, Apr 25 (4:30pm CT)
    27: ('Minnesota 3, Dallas 2 (OT)',
         'MIN: M. Boldy (OT winner, 29s left), +2 (unconfirmed) | DAL: 2 goals (scorers unconfirmed)'),
    # Match 27 (row 28): Pittsburgh @ Philadelphia, Game 4, Apr 25 (8pm ET)
    28: ('Pittsburgh 4, Philadelphia 2',
         'PIT: S. Crosby, R. Rakell, K. Letang, C. Dewar (EN) | PHI: 2 goals (scorers unconfirmed)'),
}

for row, (result, scorers) in updates.items():
    ws.cell(row=row, column=9).value = result   # I = Result
    ws.cell(row=row, column=10).value = scorers # J = Goal Scorers

# --- Update "Bracket" sheet: series status text ---
bracket = wb['Bracket']
bracket_updates = {
    'B7':  'Buffalo leads 2-1',          # Buffalo–Boston (G3 BUF won)
    'B15': 'Montreal leads 2-1',         # Tampa Bay–Montreal (G3 MTL OT)
    'B23': 'Carolina wins 4-0',          # Carolina–Ottawa (sweep)
    'B31': 'Philadelphia leads 3-1',     # Pittsburgh–Philadelphia (G4 PIT won)
    'P7':  'Colorado leads 3-0',         # Colorado–Los Angeles
    'P15': 'Series tied 2-2',            # Dallas–Minnesota (G4 MIN OT)
    'P23': 'Utah leads 2-1',             # Vegas–Utah
    'P31': 'Anaheim leads 2-1',          # Edmonton–Anaheim
}
for coord, val in bracket_updates.items():
    bracket[coord] = val

wb.save(dst)
print('Saved:', dst)
