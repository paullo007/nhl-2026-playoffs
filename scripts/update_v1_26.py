import openpyxl
import shutil
from openpyxl.styles import PatternFill, Font

REPO = '/home/user/nhl-2026-playoffs'
src = f'{REPO}/2026 NHL Playoffs_v1.25.xlsx'
dst = f'{REPO}/2026 NHL Playoffs_v1.26.xlsx'
shutil.copyfile(src, dst)

wb = openpyxl.load_workbook(dst)
ws = wb['2026 NHL Playoffs_By Dates']
b  = wb['Bracket']

# === BY DATES UPDATES ===
# Rebuild data from run-2026-05-11 (v1.26 was never pushed; rebuild from v1.25)
# + COL/MIN G4 newly confirmed today (May 12)
# + BUF/MTL G4 today (May 12, 7 PM ET) — not yet played → TBU

updates = {
    # Row 60 (M#59): BUF/MTL G2, May 8 — MTL 5, BUF 1
    # Carry-forward from run-2026-05-11 rebuild data (originally verified May 9 session)
    60: ('Montreal 5, Buffalo 1',
         'MTL: A. Newhook (2), M. Matheson (GW), A. Carrier, N. Suzuki (EN) | BUF: Z. Benson'),

    # Row 61 (M#60): VGK/ANA G3, May 8 — VGK 6, ANA 2
    61: ('Vegas 6, Anaheim 2',
         'VGK: S. Theodore, B. McNabb (SH), M. Marner (3, 1 PP, incl. GW), B. Howden (EN) | ANA: B. Sennecke, C. Kreider'),

    # Row 62 (M#61): COL/MIN G3, May 9 — MIN 5, COL 1
    62: ('Minnesota 5, Colorado 1',
         'MIN: K. Kaprizov, Q. Hughes (PP, GW), R. Hartman (PP), B. Faber, M. Boldy (EN) | COL: N. MacKinnon (PP)'),

    # Row 63 (M#62): BUF/MTL G3, May 10 — MTL 6, BUF 2
    63: ('Montreal 6, Buffalo 2',
         'MTL: A. Newhook (2, incl. EN), C. Caufield (PP, GW), Z. Bolduc, J. Slafkovsky (PP), K. Dach | BUF: T. Thompson, R. Dahlin (PP)'),

    # Row 64 (M#63): CAR/PHI G4, May 9 (played May 9, workbook had TBD Mon May11) — CAR 3, PHI 2 OT
    # CAR sweeps PHI 4-0; wins East Metropolitan Semifinal
    64: ('Carolina 3, Philadelphia 2 (OT) (CAR wins series 4-0)',
         'CAR: J. Blake (2, incl. OT winner), L. Stankoven | PHI: T. Foerster, A. Bump'),

    # Row 65 (M#64): VGK/ANA G4, May 10 (played May 10, workbook had TBD Mon May11) — ANA 4, VGK 3
    65: ('Anaheim 4, Vegas 3',
         'ANA: B. Sennecke (PP), M. Granlund, A. Killorn (PP), I. Moore (GW) | VGK: P. Dorofeyev (PP), B. Howden, T. Hertl'),

    # Row 66 (M#65): COL/MIN G4, May 11 — COL 5, MIN 2
    # Newly confirmed today from web search (ESPN, Fox9, search snippets)
    # Goal log: Yurov PP (P1 9:46), Kadri PP (P2 6:08), Colton EV (P3 6:56),
    #           Sturm EV (P3 9:15), Kelly EV/GW (P3 11:32), MacKinnon EN (P3 19:27),
    #           Nelson EN (P3 19:52)
    # GW: Kelly (COL's 3rd goal = 1 more than MIN's final total of 2) ✓
    # COL leads series 3-1
    66: ('Colorado 5, Minnesota 2',
         'COL: N. Kadri (PP), R. Colton, P. Kelly (GW), N. MacKinnon (EN), B. Nelson (EN) | MIN: D. Yurov (PP), N. Sturm'),

    # Row 67 (M#66): BUF/MTL G4, May 12 (today, 7 PM ET) — not yet played
    67: ('TBU', 'TBU'),
}

for row, (result, scorers) in updates.items():
    ws.cell(row=row, column=9).value = result
    ws.cell(row=row, column=10).value = scorers

# === BRACKET UPDATES ===
# Recomputed series states from all confirmed (non-TBU) games:
#   COL/MIN: COL leads 3-1 (G1 COL, G2 COL, G3 MIN, G4 COL)
#   VGK/ANA: Tied 2-2 (G1 VGK, G2 ANA, G3 VGK, G4 ANA)
#   BUF/MTL: MTL leads 2-1 (G1 BUF, G2 MTL, G3 MTL; G4 TBU)
#   CAR/PHI: CAR wins 4-0 (G1 CAR, G2 CAR, G3 CAR, G4 CAR)

WHITE      = PatternFill('solid', fgColor='FFFFFF')
YELLOW     = PatternFill('solid', fgColor='FFFF00')
GREEN      = PatternFill('solid', fgColor='92D050')
GREY       = PatternFill('solid', fgColor='D9D9D9')
TAN        = PatternFill('solid', fgColor='FFF2CC')

BLACK_BOLD = Font(name='Calibri', size=11, bold=True, color='000000')
GREY_TEXT  = Font(name='Calibri', size=11, bold=True, color='808080', strike=True)

# COL/MIN — COL leads 3-1
b['D11'].value = 'COL 3 - 1 MIN'          # badge stays TAN (set by prior script)
b['D7'].fill   = YELLOW                    # COL: leading (no change from v1.25, already YELLOW)
b['D15'].fill  = WHITE                     # MIN: trailing (no change from v1.25, already WHITE)

# VGK/ANA — Tied 2-2
b['D27'].value = 'VGK 2 - 2 ANA'
b['D23'].fill  = WHITE                     # VGK: tied (no change from v1.25, already WHITE)
b['D31'].fill  = WHITE                     # ANA: tied (no change from v1.25, already WHITE)

# BUF/MTL — MTL leads 2-1 (G4 TBU, do not count)
b['N11'].value = 'BUF 1 - 2 MTL'
b['N7'].fill   = WHITE                     # BUF: trailing (was YELLOW in v1.25)
b['N15'].fill  = YELLOW                    # MTL: leading (was WHITE in v1.25)
b['N7'].font   = BLACK_BOLD
b['N15'].font  = BLACK_BOLD

# CAR/PHI — CAR wins 4-0 (sweep)
b['N27'].value = 'CAR 4 - 0 PHI'
b['N23'].fill  = GREEN                     # CAR: clinched (was YELLOW in v1.25)
b['N31'].fill  = GREY                      # PHI: eliminated (was WHITE in v1.25)
b['N23'].font  = BLACK_BOLD
b['N31'].font  = GREY_TEXT                 # strikethrough for eliminated team

# CAR advances to Eastern Conference Metropolitan Final
# L27 = ECF Metropolitan side team cell
b['L27'].value = 'Carolina Hurricanes\n(advanced — opponent TBD)'
b['L27'].fill  = GREEN
b['L27'].font  = BLACK_BOLD

wb.save(dst)
print(f'Saved: {dst}')
