"""
2026-05-20 daily update: v1.33 → v1.34

Actions:
- Add Conference Final game rows (Matches 79-92) to the By Dates sheet.
  - Match 79 (WCF G1, tonight May 20 at 8 PM ET): game not yet played → mark TBU.
  - Matches 80-92 (May 21 onward): future games → Result/Scorers = None.
- Add CF series-score badges to the Bracket sheet:
  - F19: "COL 0 - 0 VGK" (WCF badge between F11/F27)
  - L19: "MTL 0 - 0 CAR" (ECF badge between L11/L27)
- No game results to confirm today (WCF G1 has not been played; ECF starts May 21).
"""
import openpyxl
import shutil
from copy import copy
from openpyxl.styles import PatternFill, Font

REPO = '/home/user/nhl-2026-playoffs'
src = f'{REPO}/2026 NHL Playoffs_v1.33.xlsx'
dst = f'{REPO}/2026 NHL Playoffs_v1.34.xlsx'
shutil.copyfile(src, dst)

wb = openpyxl.load_workbook(dst)
ws = wb['2026 NHL Playoffs_By Dates']
b  = wb['Bracket']

# ── Color / font constants (matching existing conventions) ─────────────────────
TAN        = PatternFill('solid', fgColor='FFF2CC')
BLACK_BOLD = Font(name='Calibri', size=11, bold=True, color='000000')

# ── Reference row for style copy-forward ──────────────────────────────────────
# Use the last existing row (row 79, Match 78) as the style template.
ref_row = 79
ref_styles = {}
for col in range(1, 11):
    cell = ws.cell(row=ref_row, column=col)
    ref_styles[col] = {
        'fill':          copy(cell.fill),
        'font':          copy(cell.font),
        'alignment':     copy(cell.alignment),
        'border':        copy(cell.border),
        'number_format': cell.number_format,
    }
existing_height = ws.row_dimensions[ref_row].height or 22

# ── CF game rows (chronological) ──────────────────────────────────────────────
# Columns: match, round, series, game, matchup, date_local, date_sgt, arena, result, scorers
# - Match 79 (WCF G1, May 20): today → TBU (game at 8 PM ET, not played yet)
# - Matches 80-92: future → None

new_rows = [
    # Match 79 — WCF G1: VGK at COL — Wed May20 8PM ET → TBU (game tonight, recap not yet available)
    (79, 'Conference Final', 'Colorado Avalanche vs Vegas Golden Knights', 'Game 1',
     'Vegas Golden Knights vs Colorado Avalanche (at Colorado Avalanche)',
     '6:00 PM MT, Wed, May20', '8:00 AM SGT, Thu, May21',
     'Ball Arena, Denver, CO',
     'TBU', 'TBU'),

    # Match 80 — ECF G1: MTL at CAR — Thu May21 8PM ET
    (80, 'Conference Final', 'Carolina Hurricanes vs Montreal Canadiens', 'Game 1',
     'Montreal Canadiens vs Carolina Hurricanes (at Carolina Hurricanes)',
     '8:00 PM ET, Thu, May21', '8:00 AM SGT, Fri, May22',
     'Lenovo Center, Raleigh, NC',
     None, None),

    # Match 81 — WCF G2: VGK at COL — Fri May22 8PM ET
    (81, 'Conference Final', 'Colorado Avalanche vs Vegas Golden Knights', 'Game 2',
     'Vegas Golden Knights vs Colorado Avalanche (at Colorado Avalanche)',
     '6:00 PM MT, Fri, May22', '8:00 AM SGT, Sat, May23',
     'Ball Arena, Denver, CO',
     None, None),

    # Match 82 — ECF G2: MTL at CAR — Sat May23 7PM ET
    (82, 'Conference Final', 'Carolina Hurricanes vs Montreal Canadiens', 'Game 2',
     'Montreal Canadiens vs Carolina Hurricanes (at Carolina Hurricanes)',
     '7:00 PM ET, Sat, May23', '7:00 AM SGT, Sun, May24',
     'Lenovo Center, Raleigh, NC',
     None, None),

    # Match 83 — WCF G3: COL at VGK — Sun May24 8PM ET
    (83, 'Conference Final', 'Colorado Avalanche vs Vegas Golden Knights', 'Game 3',
     'Colorado Avalanche vs Vegas Golden Knights (at Vegas Golden Knights)',
     '5:00 PM PT, Sun, May24', '8:00 AM SGT, Mon, May25',
     'T-Mobile Arena, Paradise, NV',
     None, None),

    # Match 84 — ECF G3: CAR at MTL — Mon May25 8PM ET
    (84, 'Conference Final', 'Carolina Hurricanes vs Montreal Canadiens', 'Game 3',
     'Carolina Hurricanes vs Montreal Canadiens (at Montreal Canadiens)',
     '8:00 PM ET, Mon, May25', '8:00 AM SGT, Tue, May26',
     'Bell Centre, Montreal, QC',
     None, None),

    # Match 85 — WCF G4: COL at VGK — Tue May26 9PM ET
    (85, 'Conference Final', 'Colorado Avalanche vs Vegas Golden Knights', 'Game 4',
     'Colorado Avalanche vs Vegas Golden Knights (at Vegas Golden Knights)',
     '6:00 PM PT, Tue, May26', '9:00 AM SGT, Wed, May27',
     'T-Mobile Arena, Paradise, NV',
     None, None),

    # Match 86 — ECF G4: CAR at MTL — Wed May27 8PM ET
    (86, 'Conference Final', 'Carolina Hurricanes vs Montreal Canadiens', 'Game 4',
     'Carolina Hurricanes vs Montreal Canadiens (at Montreal Canadiens)',
     '8:00 PM ET, Wed, May27', '8:00 AM SGT, Thu, May28',
     'Bell Centre, Montreal, QC',
     None, None),

    # Match 87 — WCF G5: VGK at COL — Thu May28 8PM ET
    (87, 'Conference Final', 'Colorado Avalanche vs Vegas Golden Knights', 'Game 5',
     'Vegas Golden Knights vs Colorado Avalanche (at Colorado Avalanche)',
     '6:00 PM MT, Thu, May28', '8:00 AM SGT, Fri, May29',
     'Ball Arena, Denver, CO',
     None, None),

    # Match 88 — ECF G5: MTL at CAR — Fri May29 8PM ET
    (88, 'Conference Final', 'Carolina Hurricanes vs Montreal Canadiens', 'Game 5',
     'Montreal Canadiens vs Carolina Hurricanes (at Carolina Hurricanes)',
     '8:00 PM ET, Fri, May29', '8:00 AM SGT, Sat, May30',
     'Lenovo Center, Raleigh, NC',
     None, None),

    # Match 89 — WCF G6: COL at VGK — Sat May30 8PM ET
    (89, 'Conference Final', 'Colorado Avalanche vs Vegas Golden Knights', 'Game 6',
     'Colorado Avalanche vs Vegas Golden Knights (at Vegas Golden Knights)',
     '5:00 PM PT, Sat, May30', '8:00 AM SGT, Sun, May31',
     'T-Mobile Arena, Paradise, NV',
     None, None),

    # Match 90 — ECF G6: CAR at MTL — Sun May31 TBD
    (90, 'Conference Final', 'Carolina Hurricanes vs Montreal Canadiens', 'Game 6',
     'Carolina Hurricanes vs Montreal Canadiens (at Montreal Canadiens)',
     'TBD, Sun, May31', 'TBD, Mon, Jun1',
     'Bell Centre, Montreal, QC',
     None, None),

    # Match 91 — WCF G7: VGK at COL — Mon Jun1 8PM ET
    (91, 'Conference Final', 'Colorado Avalanche vs Vegas Golden Knights', 'Game 7',
     'Vegas Golden Knights vs Colorado Avalanche (at Colorado Avalanche)',
     '6:00 PM MT, Mon, Jun1', '8:00 AM SGT, Tue, Jun2',
     'Ball Arena, Denver, CO',
     None, None),

    # Match 92 — ECF G7: MTL at CAR — Tue Jun2 8PM ET
    (92, 'Conference Final', 'Carolina Hurricanes vs Montreal Canadiens', 'Game 7',
     'Montreal Canadiens vs Carolina Hurricanes (at Carolina Hurricanes)',
     '8:00 PM ET, Tue, Jun2', '8:00 AM SGT, Wed, Jun3',
     'Lenovo Center, Raleigh, NC',
     None, None),
]

start_row = ref_row + 1  # row 80

for idx, row_data in enumerate(new_rows):
    r = start_row + idx
    match_num, rnd, series, game, matchup, dl, ds, arena, result, scorers = row_data

    ws.cell(row=r, column=1).value = match_num
    ws.cell(row=r, column=2).value = rnd
    ws.cell(row=r, column=3).value = series
    ws.cell(row=r, column=4).value = game
    ws.cell(row=r, column=5).value = matchup
    ws.cell(row=r, column=6).value = dl
    ws.cell(row=r, column=7).value = ds
    ws.cell(row=r, column=8).value = arena
    ws.cell(row=r, column=9).value = result    # TBU or None
    ws.cell(row=r, column=10).value = scorers  # TBU or None

    # Apply reference styles
    for col in range(1, 11):
        cell = ws.cell(row=r, column=col)
        s = ref_styles[col]
        cell.fill          = copy(s['fill'])
        cell.font          = copy(s['font'])
        cell.alignment     = copy(s['alignment'])
        cell.border        = copy(s['border'])
        cell.number_format = s['number_format']

    ws.row_dimensions[r].height = existing_height

# ── Bracket: add CF series-score badges ───────────────────────────────────────
# F19 = WCF badge (between F11=COL and F27=VGK): COL 0 - 0 VGK
# L19 = ECF badge (between L11=MTL and L27=CAR): MTL 0 - 0 CAR
# Both series at 0-0 (no games played; WCF G1 is TBU).

for coord, val in [('F19', 'COL 0 - 0 VGK'), ('L19', 'MTL 0 - 0 CAR')]:
    cell = b[coord]
    cell.value = val
    cell.fill  = TAN
    cell.font  = BLACK_BOLD

wb.save(dst)
print(f'Saved: {dst}')
print(f'Added {len(new_rows)} CF game rows (Matches 79-92, rows 80-93)')
print(f'Max row in By Dates: {ws.max_row}')
print('Bracket CF badges added: F19="COL 0 - 0 VGK", L19="MTL 0 - 0 CAR"')
