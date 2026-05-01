import openpyxl
import shutil
from openpyxl.styles import PatternFill, Font

REPO = '/home/user/nhl-2026-playoffs'
src = f'{REPO}/2026 NHL Playoffs_v1.18.xlsx'
dst = f'{REPO}/2026 NHL Playoffs_v1.19.xlsx'
shutil.copyfile(src, dst)

wb = openpyxl.load_workbook(dst)
ws = wb['2026 NHL Playoffs_By Dates']

# --- By Dates updates ---
# Confirmed games (Apr 30):
#   Match 40 row 41: EDM @ ANA Game 6 — ANA wins 5-2 (series 4-2)
#   Match 41 row 42: DAL @ MIN Game 6 — MIN wins 5-2 (series 4-2)
# TBU games (May 1, not yet played):
#   Match 42 row 43: BUF @ BOS Game 6
#   Match 43 row 44: TBL @ MTL Game 6
#   Match 44 row 45: VGK @ UTA Game 6

updates = {
    41: ('Anaheim 5, Edmonton 2 (ANA wins series 4-2)',
         'ANA: R. Poehling, C. Kreider, C. Gauthier (PP, GW), T. Terry, L. Carlsson (EN) | EDM: C. Murphy, V. Podkolzin'),
    42: ('Minnesota 5, Dallas 2 (MIN wins series 4-2)',
         'MIN: Q. Hughes (2, incl. GW), V. Tarasenko, M. Boldy (2, both EN) | DAL: W. Johnston (PP), M. Bourque'),
    43: ('TBU', 'TBU'),  # BUF @ BOS Game 6 — not yet played
    44: ('TBU', 'TBU'),  # TBL @ MTL Game 6 — not yet played
    45: ('TBU', 'TBU'),  # VGK @ UTA Game 6 — not yet played
}

for row, (result, scorers) in updates.items():
    ws.cell(row=row, column=9).value = result
    ws.cell(row=row, column=10).value = scorers

# --- Bracket updates ---
b = wb['Bracket']

GREEN      = PatternFill('solid', fgColor='92D050')
GREY       = PatternFill('solid', fgColor='D9D9D9')
BLACK_BOLD = Font(name='Calibri', size=11, bold=True, color='000000')
GREY_TEXT  = Font(name='Calibri', size=11, bold=True, color='808080', strike=True)

# Series badges (format: TEAM_LEFT SCORE_LEFT - SCORE_RIGHT TEAM_RIGHT)
b['B15'].value = 'DAL 2 - 4 MIN'
b['B31'].value = 'EDM 2 - 4 ANA'

# B13 (DAL) — eliminated
b['B13'].fill = GREY
b['B13'].font = GREY_TEXT

# B17 (MIN) — advanced (clinched)
b['B17'].fill = GREEN
b['B17'].font = BLACK_BOLD

# B29 (EDM) — eliminated
b['B29'].fill = GREY
b['B29'].font = GREY_TEXT

# B33 (ANA) — advanced (clinched)
b['B33'].fill = GREEN
b['B33'].font = BLACK_BOLD

# D7 (COL R2) — update to show MIN as opponent now confirmed
d7 = b['D7']
d7.value = 'Colorado Avalanche\n(advanced — vs. Minnesota Wild)'
d7.fill = GREEN
d7.font = BLACK_BOLD

# D15 (MIN R2) — MIN advances, faces COL
d15 = b['D15']
d15.value = 'Minnesota Wild\n(advanced — vs. Colorado Avalanche)'
d15.fill = GREEN
d15.font = BLACK_BOLD

# D31 (ANA R2) — ANA advances, VGK/UTA series still ongoing
d31 = b['D31']
d31.value = 'Anaheim Ducks\n(advanced — opponent TBD)'
d31.fill = GREEN
d31.font = BLACK_BOLD

wb.save(dst)
print('Saved:', dst)
