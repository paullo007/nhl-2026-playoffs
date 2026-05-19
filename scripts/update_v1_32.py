import openpyxl
import shutil
from openpyxl.styles import PatternFill, Font

REPO = '/home/user/nhl-2026-playoffs'
src = f'{REPO}/2026 NHL Playoffs_v1.32.xlsx'
dst = f'{REPO}/2026 NHL Playoffs_v1.33.xlsx'
shutil.copyfile(src, dst)

wb = openpyxl.load_workbook(dst)
ws = wb['2026 NHL Playoffs_By Dates']
b  = wb['Bracket']

# === BY DATES UPDATE — May 19 2026 run ===
#
# Row 79 (M#78): MTL @ BUF, Game 7 of East R2 Atlantic, May 18
# Retry of TBU from yesterday's run (recap available today).
# MTL wins 3-2 in OT; Newhook scores at 11:22 OT; MTL wins series 4-3.
#
# Goal log (chronological):
#   P1  4:30 — P. Danault  (MTL) EV  -> 1-0 MTL (Guhle pass deflects off Danault's skate past Luukkonen)
#   P1 ~14:30 — Z. Bolduc  (MTL) PP  -> 2-0 MTL (one-timer from right circle past Luukkonen high short side)
#   P2 13:19 — J. Greenway (BUF) EV  -> 2-1 MTL (Samuelsson shot deflected off Greenway past Dobes)
#   P3  6:27 — R. Dahlin   (BUF) EV  -> 2-2 (shot past Dobes, set up by Owen Power)
#   OT 11:22 — A. Newhook  (MTL) OT winner -> 3-2 MTL (wrist shot past Luukkonen)
# GW: Newhook (OT winner) ✓
# Count: MTL 3 (Danault + Bolduc + Newhook); BUF 2 (Greenway + Dahlin) ✓
# MTL goalie: Jakub Dobes (37 saves); BUF goalie: Ukko-Pekka Luukkonen
# Series: MTL wins 4-3; advances to East CF vs Carolina Hurricanes
#
# Sources confirmed via NHL.com headlines/video URLs:
#   "Newhook wins it in OT, Canadiens defeat Sabres in Game 7 to advance to Eastern Final"
#   NHL.com video: mtl-buf-danault-scores-goal-against-ukko-pekka-luukkonen (Danault vs Luukkonen ✓)
#   NHL.com video: mtl-buf-newhook-scores-goal-against-ukko-pekka-luukkonen (Newhook vs Luukkonen ✓)
#   NHL.com video: mtl-buf-dahlin-scores-goal-against-jakub-dobes (Dahlin vs Dobes ✓)
ws.cell(row=79, column=9).value  = 'Montreal 3, Buffalo 2 (OT) (MTL wins series 4-3)'
ws.cell(row=79, column=10).value = (
    'MTL: P. Danault, Z. Bolduc (PP), A. Newhook (OT winner) | '
    'BUF: J. Greenway, R. Dahlin'
)

# === BRACKET UPDATES ===
# MTL wins R2 4-3: MTL advances to Eastern Conference Final vs CAR

WHITE      = PatternFill('solid', fgColor='FFFFFF')
GREEN      = PatternFill('solid', fgColor='92D050')
TAN        = PatternFill('solid', fgColor='FFF2CC')
GREY       = PatternFill('solid', fgColor='D9D9D9')

BLACK_BOLD = Font(name='Calibri', size=11, bold=True, color='000000')
GREY_TEXT  = Font(name='Calibri', size=11, bold=True, color='808080', strike=True)

# N11 badge: BUF 3 - 3 MTL -> BUF 3 - 4 MTL (series final)
b['N11'].value = 'BUF 3 - 4 MTL'
b['N11'].fill  = TAN
b['N11'].font  = BLACK_BOLD

# N7 (BUF R2 cell): WHITE -> GREY with strikethrough (eliminated)
b['N7'].fill = GREY
b['N7'].font = GREY_TEXT

# N15 (MTL R2 cell): WHITE -> GREEN (clinched/advancing)
b['N15'].fill = GREEN
b['N15'].font = BLACK_BOLD

# L11: Atlantic Final placeholder -> MTL advances, facing CAR
b['L11'].value = 'Montreal Canadiens\n(advanced — vs. Carolina Hurricanes)'
b['L11'].fill  = GREEN
b['L11'].font  = BLACK_BOLD

# L27: CAR was "opponent TBD" -> update to vs. Montreal Canadiens
b['L27'].value = 'Carolina Hurricanes\n(advanced — vs. Montreal Canadiens)'
b['L27'].fill  = GREEN
b['L27'].font  = BLACK_BOLD

wb.save(dst)
print(f'Saved: {dst}')
