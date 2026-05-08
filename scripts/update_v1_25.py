import openpyxl
import shutil
from openpyxl.styles import PatternFill, Font

REPO = '/home/user/nhl-2026-playoffs'
src = f'{REPO}/2026 NHL Playoffs_v1.24.xlsx'
dst = f'{REPO}/2026 NHL Playoffs_v1.25.xlsx'
shutil.copyfile(src, dst)

wb = openpyxl.load_workbook(dst)
ws = wb['2026 NHL Playoffs_By Dates']
b  = wb['Bracket']

# === BY DATES UPDATES ===

# Row 58 (Match 57): VGK vs ANA Game 2, May 6 — CORRECTION
# Previous run had this wrong (VGK 3, ANA 1). NHL.com recap confirms ANA won 3-1.
# ANA: Sennecke (EV, P2 11:23), Carlsson (GW, P3 6:36), Harkins (EN, P3 16:30)
# VGK: Stone (PP, 6 sec left)
ws.cell(row=58, column=9).value  = 'Anaheim 3, Vegas 1'
ws.cell(row=58, column=10).value = 'ANA: B. Sennecke, L. Carlsson (GW), J. Harkins (EN) | VGK: M. Stone (PP)'

# Row 59 (Match 58): CAR vs PHI Game 3, May 7 — confirmed from NHL.com recap
# Goals: Staal PP (P1 17:27), Zegras EV (P2 2:31), Chatfield SH/GW (P2 15:59),
#         Svechnikov PP 4-on-3 (P3 3:52), Ehlers EV breakaway (P3 7:08)
# GW: Chatfield (2nd CAR goal, one more than PHI's final total of 1)
ws.cell(row=59, column=9).value  = 'Carolina 4, Philadelphia 1'
ws.cell(row=59, column=10).value = 'CAR: J. Staal (PP), J. Chatfield (SH, GW), A. Svechnikov (PP), N. Ehlers | PHI: T. Zegras'

# Row 60 (Match 59): BUF vs MTL Game 2, May 8 — tonight (7 PM ET), no recap yet
ws.cell(row=60, column=9).value  = 'TBU'
ws.cell(row=60, column=10).value = 'TBU'

# Row 61 (Match 60): VGK vs ANA Game 3, May 8 — tonight (9:30 PM ET), no recap yet
ws.cell(row=61, column=9).value  = 'TBU'
ws.cell(row=61, column=10).value = 'TBU'

# === BRACKET UPDATES ===

WHITE  = PatternFill('solid', fgColor='FFFFFF')
YELLOW = PatternFill('solid', fgColor='FFFF00')
TAN    = PatternFill('solid', fgColor='FFF2CC')

BLACK_BOLD = Font(name='Calibri', size=11, bold=True, color='000000')

# VGK/ANA series corrected to 1-1 (ANA won G2)
# Badge D27: was "VGK 2 - 0 ANA", correct to "VGK 1 - 1 ANA"
b['D27'].value = 'VGK 1 - 1 ANA'
# D23 (VGK team cell): was YELLOW (led 2-0), now WHITE (tied 1-1)
b['D23'].fill = WHITE

# CAR/PHI series now 3-0 after G3 confirmed
# Badge N27: was "CAR 2 - 0 PHI", update to "CAR 3 - 0 PHI"
b['N27'].value = 'CAR 3 - 0 PHI'
# N23 (CAR) stays YELLOW (leading 3-0, not yet clinched) — no change needed
# N31 (PHI) stays WHITE (trailing 0-3) — no change needed

wb.save(dst)
print(f'Saved: {dst}')
