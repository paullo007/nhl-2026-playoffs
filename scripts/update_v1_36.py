"""
2026-05-22 daily update: v1.35 → v1.36

Actions:
- Match 80 (ECF G1, Montreal at Carolina, May 21): MTL 6, CAR 2 — confirmed.
  Scorers: MTL: C. Caufield, P. Danault, A. Texier (GW), I. Demidov,
           J. Slafkovsky (2 incl. EN) | CAR: S. Jarvis, E. Robinson
- Match 81 (WCF G2, Vegas at Colorado, May 22): game tonight 8 PM ET → TBU.
- Bracket: ECF badge L19 MTL 0-0 CAR → MTL 1-0 CAR.
           L11 (MTL) fill GREEN → YELLOW (leading 1-0).
           L27 (CAR) fill GREEN → WHITE (trailing 0-1).
"""
import openpyxl
import shutil
from openpyxl.styles import PatternFill, Font

REPO = '/home/user/nhl-2026-playoffs'
src = f'{REPO}/2026 NHL Playoffs_v1.35.xlsx'
dst = f'{REPO}/2026 NHL Playoffs_v1.36.xlsx'
shutil.copyfile(src, dst)

wb = openpyxl.load_workbook(dst)
ws = wb['2026 NHL Playoffs_By Dates']
b  = wb['Bracket']

# Color / font constants
WHITE      = PatternFill('solid', fgColor='FFFFFF')
YELLOW     = PatternFill('solid', fgColor='FFFF00')
TAN        = PatternFill('solid', fgColor='FFF2CC')
BLACK_BOLD = Font(name='Calibri', size=11, bold=True, color='000000')

# ── By Dates: Row 81 = Match 80 (ECF G1, MTL at CAR, May 21) ─────────────────
ws.cell(row=81, column=9).value  = 'Montreal 6, Carolina 2'
ws.cell(row=81, column=10).value = (
    'MTL: C. Caufield, P. Danault, A. Texier (GW), I. Demidov, '
    'J. Slafkovsky (2, incl. EN) | CAR: S. Jarvis, E. Robinson'
)

# ── By Dates: Row 82 = Match 81 (WCF G2, VGK at COL, May 22 — tonight) ──────
ws.cell(row=82, column=9).value  = 'TBU'
ws.cell(row=82, column=10).value = 'TBU'

# ── Bracket: ECF badge (L19) ─────────────────────────────────────────────────
b['L19'].value = 'MTL 1 - 0 CAR'
b['L19'].fill  = TAN
b['L19'].font  = BLACK_BOLD

# ── Bracket: ECF team cells ───────────────────────────────────────────────────
# L11 = Montreal Canadiens: leads series 1-0 → YELLOW
b['L11'].fill = YELLOW
b['L11'].font = BLACK_BOLD

# L27 = Carolina Hurricanes: trails series 0-1 → WHITE
b['L27'].fill = WHITE
b['L27'].font = BLACK_BOLD

wb.save(dst)
print(f'Saved: {dst}')
print('By Dates row 81 (Match 80 ECF G1): Montreal 6, Carolina 2')
print('By Dates row 82 (Match 81 WCF G2): TBU')
print('Bracket L19: MTL 1 - 0 CAR')
print('Bracket L11 (MTL): fill → YELLOW (leading 1-0)')
print('Bracket L27 (CAR): fill → WHITE (trailing 0-1)')
