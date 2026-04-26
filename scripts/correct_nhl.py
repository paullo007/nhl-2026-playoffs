import openpyxl
import shutil

src = '/Users/paullo/01_PLO/02_CLAUDE CODE/02_NHL 2026 Playoffs/2026 NHL Playoffs_v1.9.xlsx'
dst = '/Users/paullo/01_PLO/02_CLAUDE CODE/02_NHL 2026 Playoffs/2026 NHL Playoffs_v1.10.xlsx'
shutil.copyfile(src, dst)

wb = openpyxl.load_workbook(dst)
ws = wb['2026 NHL Playoffs_By Dates']

# Verified scorers per NHL.com authoritative recaps.
# Format kept consistent with existing file: "TEAM: scorer (tags), ... | TEAM: scorer (tags), ..."
updates = {
    # Row : (Result text [unchanged unless noted], Goal Scorers)
    # --- Pre-existing rows: corrections to scorers ---
    3:  ('Minnesota 6, Dallas 1',
         'MIN: J. Eriksson Ek (2, both PP, incl. GW), K. Kaprizov, R. Hartman, M. Boldy (2, incl. EN) | DAL: J. Robertson (PP)'),
    5:  ('Montreal 4, Tampa Bay 3 (OT)',
         'MTL: J. Anderson, J. Slafkovsky (3, all PP, incl. OT winner) | TBL: D. Raddysh (PP), B. Hagel (2)'),
    6:  ('Buffalo 4, Boston 3',
         'BUF: T. Thompson (2), M. Samuelsson, A. Tuch (EN) | BOS: M. Geekie, E. Lindholm, D. Pastrnak (PP)'),
    7:  ('Vegas 4, Utah 2',
         'VGK: C. Sissons, M. Stone (PP), N. Dowd, I. Barbashev (EN) | UTA: L. Cooley, K. Stenlund'),
    8:  ('Colorado 2, Los Angeles 1',
         "COL: A. Lehkonen, L. O'Connor | LAK: A. Panarin (PP)"),
    9:  ('Philadelphia 3, Pittsburgh 0',
         'PHI: P. Martone, G. Hathaway (SH), L. Glendening (EN) | PIT: (shutout)'),
    10: ('Carolina 3, Ottawa 2 (2OT)',
         'CAR: L. Stankoven (PP), S. Aho, J. Martinook (2OT winner) | OTT: D. Batherson, D. Cozens'),
    11: ('Dallas 4, Minnesota 2',
         'DAL: W. Johnston (2, incl. EN PP), M. Duchene (PP), J. Robertson | MIN: B. Faber (2)'),
    12: ('Edmonton 4, Anaheim 3',
         'EDM: J. Dickinson (2), K. Kapanen (2, incl. GW) | ANA: T. Terry (2, 1 PP), L. Carlsson'),
    13: ('Tampa Bay 3, Montreal 2 (OT)',
         'TBL: B. Hagel, N. Kucherov, J.J. Moser (OT winner) | MTL: J. Anderson, L. Hutson'),
    14: ('Boston 4, Buffalo 2',
         'BOS: V. Arvidsson (2), M. Geekie, P. Zacha (PP) | BUF: B. Byram, P. Krebs'),
    15: ('Utah 3, Vegas 2',
         'UTA: M. Weegar, D. Guenther, L. Cooley (GW) | VGK: M. Stone (PP), I. Barbashev'),
    16: ('Colorado 2, Los Angeles 1 (OT)',
         'COL: G. Landeskog, N. Roy (OT winner) | LAK: A. Panarin (PP)'),

    # --- New rows added in v1.9: refined with verified tags & previously unconfirmed scorers ---
    17: ('Philadelphia 5, Pittsburgh 2',
         'PHI: T. Zegras (PP), R. Ristolainen, N. Seeler (GW), N. Cates (PP), O. Tippett (EN) | PIT: E. Malkin (PP), E. Karlsson (PP)'),
    18: ('Dallas 4, Minnesota 3 (2OT)',
         'DAL: M. Rantanen (PP), J. Robertson, M. Duchene (PP), W. Johnston (2OT winner, PP) | MIN: M. Johansson (PP), J. Eriksson Ek, M. McCarron'),
    19: ('Anaheim 6, Edmonton 4',
         'ANA: C. Gauthier (2, 1 PP, incl. GW), R. Poehling (2, 1 SH, 1 EN), J. Trouba, A. Killorn (PP) | EDM: L. Draisaitl, C. Murphy, Z. Hyman, J. Samanski'),
    20: ('Buffalo 3, Boston 1',
         'BUF: B. Byram, A. Tuch (GW), N. Ostlund (EN) | BOS: T. Jeannot'),
    21: ('Carolina 2, Ottawa 1',
         'CAR: L. Stankoven, J. Blake (GW) | OTT: D. Batherson'),
    22: ('Colorado 4, Los Angeles 2',
         'COL: G. Landeskog, C. Makar (GW), A. Lehkonen (SH), B. Nelson (EN) | LAK: T. Moore, A. Kempe (PP)'),
    23: ('Montreal 3, Tampa Bay 2 (OT)',
         'MTL: A. Texier, K. Dach, L. Hutson (OT winner) | TBL: B. Hagel, B. Point'),
    24: ('Utah 4, Vegas 2',
         'UTA: M. Weegar, D. Guenther (PP), L. Crouse (2) | VGK: J. Eichel, N. Dowd'),
    25: ('Anaheim 7, Edmonton 4',
         'ANA: M. McTavish, M. Granlund (PP), A. Killorn, B. Sennecke, L. Carlsson, J. Viel, J. LaCombe (EN) | EDM: V. Podkolzin, K. Kapanen, R. Nugent-Hopkins, C. McDavid (PP)'),
    26: ('Carolina 4, Ottawa 2 (CAR wins series 4-0)',
         'CAR: T. Hall, L. Stankoven (GW), S. Aho (2, both EN) | OTT: D. Batherson (PP), D. Cozens'),
    27: ('Minnesota 3, Dallas 2 (OT)',
         'MIN: B. Faber, M. Foligno, M. Boldy (OT winner) | DAL: J. Robertson (PP), M. Heiskanen (PP)'),
    28: ('Pittsburgh 4, Philadelphia 2',
         'PIT: S. Crosby (PP), R. Rakell, K. Letang, C. Dewar (EN) | PHI: D. Barkey, T. Konecny'),
}

for row, (result, scorers) in updates.items():
    ws.cell(row=row, column=9).value = result
    ws.cell(row=row, column=10).value = scorers

wb.save(dst)
print('Saved:', dst)
