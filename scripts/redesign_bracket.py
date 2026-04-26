import openpyxl
import shutil
from openpyxl.styles import PatternFill, Font, Alignment

src = '/Users/paullo/01_PLO/02_CLAUDE CODE/02_NHL 2026 Playoffs/2026 NHL Playoffs_v1.10.xlsx'
dst = '/Users/paullo/01_PLO/02_CLAUDE CODE/02_NHL 2026 Playoffs/2026 NHL Playoffs_v1.11.xlsx'
shutil.copyfile(src, dst)

wb = openpyxl.load_workbook(dst)
ws = wb['Bracket']

# Color palette (matching existing file conventions)
WHITE   = PatternFill('solid', fgColor='FFFFFF')
GREEN   = PatternFill('solid', fgColor='92D050')   # Advanced / clinched
YELLOW  = PatternFill('solid', fgColor='FFFF00')   # Currently leading
TAN     = PatternFill('solid', fgColor='FFF2CC')   # Series score badge
GREY    = PatternFill('solid', fgColor='D9D9D9')   # Eliminated

BLACK_BOLD = Font(name='Calibri', size=11, bold=True, color='000000')
BLACK_NORMAL = Font(name='Calibri', size=11, bold=False, color='000000')
GREY_TEXT  = Font(name='Calibri', size=11, bold=True, color='808080', strike=True)
RED_BOLD = Font(name='Calibri', size=11, bold=True, color='C00000')

# === LEFT SIDE (Western Conference, top to bottom: COL/LAK, DAL/MIN, VGK/UTA, EDM/ANA) ===

# Header
ws['A3'] = 'WESTERN CONFERENCE'

# Round 1 left — team boxes + series score badges
# (cell, value, fill, font)
left_team_data = [
    ('B5',  'Colorado Avalanche\n(CEN #1 · 121 pts)',     YELLOW,  BLACK_BOLD),  # leads 3-0
    ('B7',  'COL 3 / LAK 0',                              TAN,     BLACK_BOLD),
    ('B9',  'Los Angeles Kings\n(PAC WC2 · 90 pts)',      WHITE,   BLACK_BOLD),

    ('B13', 'Dallas Stars\n(CEN #2 · 112 pts)',           WHITE,   BLACK_BOLD),
    ('B15', 'DAL 2 / MIN 2',                              TAN,     BLACK_BOLD),
    ('B17', 'Minnesota Wild\n(CEN #3 · 104 pts)',         WHITE,   BLACK_BOLD),

    ('B21', 'Vegas Golden Knights\n(PAC #1 · 95 pts)',    WHITE,   BLACK_BOLD),
    ('B23', 'VGK 1 / UTA 2',                              TAN,     BLACK_BOLD),
    ('B25', 'Utah Mammoth\n(CEN WC1 · 92 pts)',           YELLOW,  BLACK_BOLD),  # leads 2-1

    ('B29', 'Edmonton Oilers\n(PAC #2 · 93 pts)',         WHITE,   BLACK_BOLD),
    ('B31', 'EDM 1 / ANA 2',                              TAN,     BLACK_BOLD),
    ('B33', 'Anaheim Ducks\n(PAC #3 · 92 pts)',           YELLOW,  BLACK_BOLD),  # leads 2-1
]
for coord, val, fill, font in left_team_data:
    c = ws[coord]
    c.value = val
    c.fill = fill
    c.font = font

# Round 2 left (Western)
ws['D7']  = 'S5 Winner\n(COL / LAK)'
ws['D15'] = 'S6 Winner\n(DAL / MIN)'
ws['D23'] = 'S7 Winner\n(VGK / UTA)'
ws['D31'] = 'S8 Winner\n(EDM / ANA)'

# Conference Final left (Western)
ws['F11'] = 'Central Final\nWinner of S5 / S6'
ws['F27'] = 'Pacific Final\nWinner of S7 / S8'

# === RIGHT SIDE (Eastern Conference, top to bottom: BUF/BOS, TBL/MTL, CAR/OTT, PIT/PHI) ===

# Header
ws['K3'] = 'EASTERN CONFERENCE'
ws['K3'].fill = WHITE  # was previously green; reset to white now that East is on the right

right_team_data = [
    ('P5',  'Buffalo Sabres\n(ATL #1 · 109 pts)',         YELLOW,  BLACK_BOLD),  # leads 2-1
    ('P7',  'BUF 2 / BOS 1',                              TAN,     BLACK_BOLD),
    ('P9',  'Boston Bruins\n(ATL WC1 · 100 pts)',         WHITE,   BLACK_BOLD),

    ('P13', 'Tampa Bay Lightning\n(ATL #2 · 106 pts)',    WHITE,   BLACK_BOLD),
    ('P15', 'TBL 1 / MTL 2',                              TAN,     BLACK_BOLD),
    ('P17', 'Montreal Canadiens\n(ATL #3 · 106 pts)',     YELLOW,  BLACK_BOLD),  # leads 2-1

    ('P21', 'Carolina Hurricanes\n(MET #1 · 113 pts)',    GREEN,   BLACK_BOLD),  # advanced
    ('P23', 'CAR 4 / OTT 0',                              TAN,     BLACK_BOLD),
    ('P25', 'Ottawa Senators\n(MET WC2 · 99 pts)',        GREY,    GREY_TEXT),   # eliminated

    ('P29', 'Pittsburgh Penguins\n(MET #2 · 98 pts)',     WHITE,   BLACK_BOLD),
    ('P31', 'PIT 1 / PHI 3',                              TAN,     BLACK_BOLD),
    ('P33', 'Philadelphia Flyers\n(MET #3 · 98 pts)',     YELLOW,  BLACK_BOLD),  # leads 3-1
]
for coord, val, fill, font in right_team_data:
    c = ws[coord]
    c.value = val
    c.fill = fill
    c.font = font

# Round 2 right (Eastern) — CAR has clinched, awaiting opponent
ws['N7']  = 'S1 Winner\n(BUF / BOS)'
ws['N15'] = 'S2 Winner\n(TBL / MTL)'
ws['N23'] = 'Carolina Hurricanes\n(advanced — opponent TBD)'
ws['N23'].fill = GREEN
ws['N23'].font = BLACK_BOLD
ws['N31'] = 'S4 Winner\n(PIT / PHI)'

# Conference Final right (Eastern)
ws['L11'] = 'Atlantic Final\nWinner of S1 / S2'
ws['L27'] = 'Metropolitan Final\nWinner of S3 / S4'

# === CENTER (Stanley Cup Final + Champion labels) — swap labels to match sides ===
ws['H4'] = 'WESTERN CHAMPION'
ws['J4'] = 'EASTERN CHAMPION'
ws['H19'] = 'WESTERN\nCONFERENCE\nCHAMPION'
ws['J19'] = 'EASTERN\nCONFERENCE\nCHAMPION'

# Source footer
ws['A38'] = 'Source: NHL.com 2026 Playoffs Bracket — https://www.nhl.com/playoffs/2026/bracket'

wb.save(dst)
print('Saved:', dst)
