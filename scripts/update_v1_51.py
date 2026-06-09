import openpyxl
import shutil

src = '/home/user/nhl-2026-playoffs/2026 NHL Playoffs_v1.50.xlsx'
dst = '/home/user/nhl-2026-playoffs/2026 NHL Playoffs_v1.51.xlsx'
shutil.copyfile(src, dst)

wb = openpyxl.load_workbook(dst)
ws = wb['2026 NHL Playoffs_By Dates']

# Row 97: Match 96, SCF Game 4, CAR @ VGK, Jun 9, 2026
# Game scheduled 5:00 PM PT / 8:00 PM ET tonight.
# NHL.com recap returns 403; WebSearch finds only preview articles — game not yet played.
# Mark TBU for retry in next run.
ws.cell(row=97, column=9).value = 'TBU'
ws.cell(row=97, column=10).value = 'TBU'

# Bracket: no changes — TBU rows do not count toward series scores.
# Current state (carried from v1.50): VGK 2 - 1 CAR (I18), VGK leads YELLOW (H19), CAR WHITE (J19).

wb.save(dst)
print(f'Saved: {dst}')
print(f'  Row 97 col I: {ws.cell(row=97, column=9).value}')
print(f'  Row 97 col J: {ws.cell(row=97, column=10).value}')
