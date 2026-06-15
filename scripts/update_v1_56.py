import openpyxl
import shutil
from openpyxl.styles import PatternFill, Font

src = '/home/user/nhl-2026-playoffs/2026 NHL Playoffs_v1.55.xlsx'
dst = '/home/user/nhl-2026-playoffs/2026 NHL Playoffs_v1.56.xlsx'
shutil.copyfile(src, dst)

wb = openpyxl.load_workbook(dst)
ws = wb['2026 NHL Playoffs_By Dates']

# Row 99: Match#98, SCF Game 6, CAR @ VGK, June 14 2026
# CAR wins 3-0, clinches Stanley Cup 4-2
# Sources: multiple web search results (NHL.com blocked with 403)
# Scorers confirmed by TSN, CBS Sports, WRAL, Yahoo Sports snippets
ws.cell(row=99, column=9).value = 'Carolina 3, Vegas 0 (CAR wins Stanley Cup 4-2)'
ws.cell(row=99, column=10).value = 'CAR: T. Hall (GW), J. Blake, N. Ehlers (EN)'

# --- Update Bracket sheet ---
b = wb['Bracket']

GREEN  = PatternFill('solid', fgColor='92D050')
GREY   = PatternFill('solid', fgColor='D9D9D9')

BLACK_BOLD = Font(name='Calibri', size=11, bold=True, color='000000')
GREY_TEXT  = Font(name='Calibri', size=11, bold=True, color='808080', strike=True)

# SCF badge: series is over VGK 2 - 4 CAR
b['I18'].value = 'STANLEY\nCUP\nFINAL\nVGK 2 - 4 CAR'

# H19 = Vegas Golden Knights (Western Champion, eliminated in SCF)
b['H19'].value = 'Vegas Golden Knights\n(lost SCF 2-4)'
b['H19'].fill = GREY
b['H19'].font = GREY_TEXT

# J19 = Carolina Hurricanes (Eastern Champion, 2026 Stanley Cup Champions)
b['J19'].value = 'Carolina Hurricanes\n(2026 Stanley Cup Champions)'
b['J19'].fill = GREEN
b['J19'].font = BLACK_BOLD

wb.save(dst)
print('Saved:', dst)
print()
print('Updated cells:')
print('  By Dates row 99 col I:', ws.cell(row=99, column=9).value)
print('  By Dates row 99 col J:', ws.cell(row=99, column=10).value)
print('  Bracket I18:', b['I18'].value)
print('  Bracket H19:', b['H19'].value)
print('  Bracket J19:', b['J19'].value)
