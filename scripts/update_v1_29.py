import openpyxl
import shutil
from openpyxl.styles import PatternFill, Font

REPO = '/home/user/nhl-2026-playoffs'
src = f'{REPO}/2026 NHL Playoffs_v1.28.xlsx'
dst = f'{REPO}/2026 NHL Playoffs_v1.29.xlsx'
shutil.copyfile(src, dst)

wb = openpyxl.load_workbook(dst)
ws = wb['2026 NHL Playoffs_By Dates']
b  = wb['Bracket']

# === BY DATES UPDATE — May 15 2026 run ===
#
# Row 71 (M#70): MTL @ BUF, Game 5, May 14 — was TBU, now confirmed
# Final: MTL 6, BUF 3
# Goal log (chronological):
#   P1  2:00 — J. Zucker (BUF) EV -> 1-0 BUF (deflection off J. Quinn's wrist shot)
#   P1  6:31 — C. Caufield (MTL) EV -> 1-1 (redirected N. Suzuki pass)
#   P1  7:45 — J. Doan (BUF) EV -> 2-1 BUF (one-timer from above left circle)
#   P1  7:54 — A. Texier (MTL) EV -> 2-2 (deflection, 9 seconds after Doan)
#   P1 10:15 — K. Helenius (BUF) EV -> 3-2 BUF (wrist shot from right circle, through legs; 1st playoff goal)
#              BUF "rattled off three goals on four shots in first 10 minutes" — P1 ends BUF 3, MTL 2
#   P2  8:01 — J. Anderson (MTL) EV -> 3-3
#   P2 16:15 — J. Evans (MTL) EV GW -> 4-3 MTL ("swept loose puck, 3:45 remaining in P2"; MTL never led until this goal)
#   P2 17:33 — N. Suzuki (MTL) PP -> 5-3 MTL
#   P3        — I. Demidov (MTL) EV -> 6-3 MTL (first career playoff goal)
# GW: Evans — MTL's 4th goal (BUF final = 3; 4 > 3) ✓
# Count: MTL 6 (Caufield + Texier + Anderson + Evans + Suzuki + Demidov); BUF 3 (Zucker + Doan + Helenius) ✓
# Series: MTL leads 3-2
ws.cell(row=71, column=9).value  = 'Montreal 6, Buffalo 3'
ws.cell(row=71, column=10).value = 'MTL: C. Caufield, A. Texier, J. Anderson, J. Evans (GW), N. Suzuki (PP), I. Demidov | BUF: J. Zucker, J. Doan, K. Helenius'

# === BRACKET UPDATES ===
# BUF/MTL: MTL wins G5 6-3 -> MTL leads series 3-2
# Badge N11: BUF 2 - 2 MTL -> BUF 2 - 3 MTL
# N7 (BUF team cell): stays WHITE (trailing 2-3)
# N15 (MTL team cell): WHITE -> YELLOW (leading 3-2)

WHITE      = PatternFill('solid', fgColor='FFFFFF')
YELLOW     = PatternFill('solid', fgColor='FFFF00')
BLACK_BOLD = Font(name='Calibri', size=11, bold=True, color='000000')

b['N11'].value = 'BUF 2 - 3 MTL'
b['N7'].fill   = WHITE       # BUF: trailing (no change)
b['N7'].font   = BLACK_BOLD
b['N15'].fill  = YELLOW      # MTL: leading (was WHITE, now YELLOW)
b['N15'].font  = BLACK_BOLD

wb.save(dst)
print(f'Saved: {dst}')
