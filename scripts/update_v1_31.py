import openpyxl
import shutil
from openpyxl.styles import PatternFill, Font

REPO = '/home/user/nhl-2026-playoffs'
src = f'{REPO}/2026 NHL Playoffs_v1.30.xlsx'
dst = f'{REPO}/2026 NHL Playoffs_v1.31.xlsx'
shutil.copyfile(src, dst)

wb = openpyxl.load_workbook(dst)
ws = wb['2026 NHL Playoffs_By Dates']
b  = wb['Bracket']

# === BY DATES UPDATE — May 17 2026 run ===
#
# Row 75 (M#74): BUF @ MTL, Game 6 of East R2 Atlantic, May 16
# BUF wins 8-3 to tie series 3-3; forces Game 7 on May 18 at Buffalo
#
# Goal log (chronological):
#   P1  0:32 — R. Dahlin    (BUF) EV  -> 1-0 BUF
#   P1  1:40 — A. Xhekaj   (MTL) EV  -> 1-1
#   P1  8:12 — I. Demidov  (MTL) PP  -> 2-1 MTL (one-timer off Hutson pass)
#   P1 10:14 — J. Evans    (MTL) SH  -> 3-1 MTL (short-handed break; BUF pulls Lyon, UPL in)
#   P1 13:56 — J. Zucker   (BUF) PP  -> 3-2 BUF (against Dobes)
#   P2  1:00 — Z. Benson   (BUF) EV  -> 3-3 (backhand rebound from behind net)
#   P2 10:54 — J. Quinn    (BUF) PP  -> 4-3 BUF (GW — one-timer right point; BUF lead held to final)
#   P2 12:59 — K. Helenius (BUF) EV  -> 5-3 BUF (2-on-1 one-timer, feed from Zucker)
#   P3  9:58 — J. Quinn    (BUF) PP  -> 6-3 BUF
#   P3 14:12 — T. Thompson (BUF) EN  -> 7-3 BUF
#   P3 17:47 — Z. Metsa    (BUF) PP  -> 8-3 BUF (against Jacob Fowler, MTL backup)
# GW: Quinn's first goal (P2, 10:54 PP) → 4-3, BUF never trailed again ✓
# Count: BUF 8 (Dahlin + Zucker + Benson + Quinn×2 + Helenius + Thompson + Metsa);
#        MTL 3 (Xhekaj + Demidov + Evans) ✓
# BUF PP: 4-for-6 (Zucker, Quinn×2, Metsa)
# Rasmus Dahlin: 1G 4A = 5 points; Tage Thompson: 1G 3A = 4 points
# BUF goalie: Alex Lyon (pulled at 10:14 P1 after Evans SH goal); UPL (Ukko-Pekka Luukkonen) 18 saves
# MTL goalie: Jakub Dobes (pulled mid-game); Jacob Fowler (backup, faced final goals)
# Series: tied 3-3; Game 7 Monday May 18 at KeyBank Center, Buffalo (7:30 PM ET)
ws.cell(row=75, column=9).value  = 'Buffalo 8, Montreal 3'
ws.cell(row=75, column=10).value = (
    'BUF: R. Dahlin, J. Zucker (PP), Z. Benson, J. Quinn (2, both PP, incl. GW), '
    'K. Helenius, T. Thompson (EN), Z. Metsa (PP) | '
    'MTL: A. Xhekaj, I. Demidov (PP), J. Evans (SH)'
)

# === BRACKET UPDATES ===
# BUF ties series 3-3: MTL no longer leads → MTL cell back to WHITE (tied)

WHITE     = PatternFill('solid', fgColor='FFFFFF')
YELLOW    = PatternFill('solid', fgColor='FFFF00')
TAN       = PatternFill('solid', fgColor='FFF2CC')

BLACK_BOLD = Font(name='Calibri', size=11, bold=True, color='000000')

# N11 badge: BUF 2 - 3 MTL -> BUF 3 - 3 MTL (series tied)
b['N11'].value = 'BUF 3 - 3 MTL'
b['N11'].fill  = TAN
b['N11'].font  = BLACK_BOLD

# N15 (MTL R2 cell): YELLOW (was leading 3-2) -> WHITE (now tied 3-3)
b['N15'].fill = WHITE
b['N15'].font = BLACK_BOLD

# N7 (BUF R2 cell): already WHITE — no change needed, but reinforce
b['N7'].fill = WHITE
b['N7'].font = BLACK_BOLD

wb.save(dst)
print(f'Saved: {dst}')
