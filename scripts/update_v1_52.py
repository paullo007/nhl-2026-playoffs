import openpyxl
import shutil
from openpyxl.styles import PatternFill, Font

src = '/home/user/nhl-2026-playoffs/2026 NHL Playoffs_v1.51.xlsx'
dst = '/home/user/nhl-2026-playoffs/2026 NHL Playoffs_v1.52.xlsx'
shutil.copyfile(src, dst)

wb = openpyxl.load_workbook(dst)
ws = wb['2026 NHL Playoffs_By Dates']

# === Row 97: Match 96, SCF Game 4, CAR @ VGK, Jun 9, 2026 ===
# Result: Carolina 5, Vegas 3 — series tied 2-2
# CAR goals: Stankoven (P1 1:06), Blake (P1 3:28), Staal PP (P1 12:48),
#             Staal GW (P3 6:32), Ehlers EN (P3 final minute)
# VGK goals: Stone (P1 7:22 breakaway), Karlsson (P2), Howden (P2 17:08)
# Sources: NHL.com recap headline + snippets, multiple aggregators; NHL.com returned 403 to direct fetch
ws.cell(row=97, column=9).value = 'Carolina 5, Vegas 3'
ws.cell(row=97, column=10).value = (
    'CAR: L. Stankoven, J. Blake, J. Staal (PP), J. Staal (GW), N. Ehlers (EN) | '
    'VGK: M. Stone, W. Karlsson, B. Howden'
)

# === Bracket updates ===
b = wb['Bracket']

WHITE     = PatternFill('solid', fgColor='FFFFFF')
BLACK_BOLD = Font(name='Calibri', size=11, bold=True, color='000000')

# SCF series badge: now tied 2-2
b['I18'].value = 'STANLEY\nCUP\nFINAL\nVGK 2 - 2 CAR'

# VGK team cell (H19): tied — WHITE (neither leading)
b['H19'].value = 'Vegas Golden Knights\n(tied SCF 2-2)'
b['H19'].fill = WHITE
b['H19'].font = BLACK_BOLD

# CAR team cell (J19): tied — WHITE
b['J19'].value = 'Carolina Hurricanes\n(tied SCF 2-2)'
b['J19'].fill = WHITE
b['J19'].font = BLACK_BOLD

wb.save(dst)
print(f'Saved: {dst}')
print('\nUpdated cells:')
print(f'  Row 97 col I: {ws.cell(row=97, column=9).value}')
print(f'  Row 97 col J: {ws.cell(row=97, column=10).value}')
print(f'  Bracket I18:  {b["I18"].value!r}')
print(f'  Bracket H19:  {b["H19"].value!r}')
print(f'  Bracket J19:  {b["J19"].value!r}')
