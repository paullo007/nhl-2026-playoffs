import openpyxl
import shutil
from openpyxl.styles import PatternFill, Font

REPO = '/home/user/nhl-2026-playoffs'
src = f'{REPO}/2026 NHL Playoffs_v1.26.xlsx'
dst = f'{REPO}/2026 NHL Playoffs_v1.27.xlsx'
shutil.copyfile(src, dst)

wb = openpyxl.load_workbook(dst)
ws = wb['2026 NHL Playoffs_By Dates']
b  = wb['Bracket']

# === BY DATES UPDATE ===
# Row 67 (M#66): BUF/MTL G4, May 12 — was TBU, now confirmed
# Final: BUF 3, MTL 2
# Goals (chronological):
#   P1 6:32 — M. Samuelsson (BUF) EV -> 1-0 BUF
#   P1 10:09 — A. Newhook (MTL) EV -> 1-1 (Jake Evans assist)
#   P1 19:47 — C. Caufield (MTL) PP -> 1-2 MTL (Slafkovsky assist)
#   P2 7:00 — T. Thompson (BUF) PP -> 2-2
#   P3 4:41 — Z. Benson (BUF) PP GW -> 3-2 BUF (Doan assist)
# GW: Benson (BUF's 3rd goal = 1 more than MTL's final total of 2)
# Count: BUF 3 (Samuelsson + Thompson + Benson) = 3; MTL 2 (Newhook + Caufield) = 2
# Note: Jack Quinn "goal" at P1 8:02 overturned — MTL coach's challenge (goalie interference by Helenius)
# Series: now tied 2-2
ws.cell(row=67, column=9).value  = 'Buffalo 3, Montreal 2'
ws.cell(row=67, column=10).value = 'BUF: M. Samuelsson, T. Thompson (PP), Z. Benson (PP, GW) | MTL: A. Newhook, C. Caufield (PP)'

# === BRACKET UPDATES ===
# BUF/MTL: BUF wins G4 3-2 -> series tied 2-2
# Badge N11: BUF 1 - 2 MTL -> BUF 2 - 2 MTL
# N7 (BUF team cell): WHITE (was WHITE trailing, stays WHITE tied)
# N15 (MTL team cell): YELLOW -> WHITE (no longer leading)

WHITE  = PatternFill('solid', fgColor='FFFFFF')
YELLOW = PatternFill('solid', fgColor='FFFF00')

BLACK_BOLD = Font(name='Calibri', size=11, bold=True, color='000000')

b['N11'].value = 'BUF 2 - 2 MTL'
b['N7'].fill   = WHITE    # BUF: tied (no change from WHITE)
b['N7'].font   = BLACK_BOLD
b['N15'].fill  = WHITE    # MTL: was YELLOW (leading), now WHITE (tied)
b['N15'].font  = BLACK_BOLD

wb.save(dst)
print(f'Saved: {dst}')
