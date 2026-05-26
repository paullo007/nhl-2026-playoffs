import openpyxl
import shutil
from openpyxl.styles import PatternFill, Font

REPO = '/home/user/nhl-2026-playoffs'
src = f'{REPO}/2026 NHL Playoffs_v1.39.xlsx'
dst = f'{REPO}/2026 NHL Playoffs_v1.40.xlsx'
shutil.copyfile(src, dst)

wb = openpyxl.load_workbook(dst)
ws = wb['2026 NHL Playoffs_By Dates']

# --- By Dates sheet updates ---
updates = {
    # Row 85: Match 84 · ECF Game 3 · CAR @ MTL · May 25, 2026
    # CAR 3, MTL 2 (OT) — CAR leads series 2-1
    # OT winner: Andrei Svechnikov (officially credited, originally awarded to Aho but corrected to Svechnikov)
    # CAR: Gostisbehere (1st EV), Hall (1st EV), Svechnikov (OT winner)
    # MTL: Matheson (1st EV), Hutson (2nd PP)
    85: (
        'Carolina 3, Montreal 2 (OT)',
        'CAR: S. Gostisbehere, T. Hall, A. Svechnikov (OT winner) | MTL: M. Matheson, L. Hutson (PP)'
    ),

    # Row 86: Match 85 · WCF Game 4 · COL @ VGK · May 26, 2026
    # Game scheduled 6:00 PM PT tonight — not yet played at time of run
    86: (
        'TBU',
        'TBU'
    ),
}

for row_num, (result, scorers) in updates.items():
    ws.cell(row=row_num, column=9).value = result
    ws.cell(row=row_num, column=10).value = scorers

# --- Bracket sheet updates ---
YELLOW = PatternFill('solid', fgColor='FFFF00')
WHITE  = PatternFill('solid', fgColor='FFFFFF')
TAN    = PatternFill('solid', fgColor='FFF2CC')

b = wb['Bracket']

# ECF series badge: was MTL 1-1 CAR, now MTL 1-2 CAR (CAR leads 2-1)
b['L19'].value = 'MTL 1 - 2 CAR'

# L27 (CAR): was WHITE (tied 1-1), now YELLOW (leading 2-1)
b['L27'].fill = YELLOW

# L11 (MTL): remains WHITE (trailing 1-2) — no change needed

# WCF: no change (COL 0-3 VGK — G4 not yet played)

wb.save(dst)
print(f'Saved: {dst}')
print()
print('By Dates updates:')
print('  Row 85: Carolina 3, Montreal 2 (OT)')
print('  Row 86: TBU (WCF G4 tonight)')
print()
print('Bracket updates:')
print('  L19: MTL 1 - 2 CAR')
print('  L27: YELLOW fill (CAR leading 2-1)')
